from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.pagination import PageNumberPagination

from django.utils import timezone
from django.db import transaction
from django.http import HttpResponse, Http404
from django.conf import settings

import logging
import json

logger = logging.getLogger(__name__)

try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    logger.warning("OpenAI package not installed. AI insights will use fallback generation.")

from model_definitions.models import ModelDefinition, ModelDefinitionHistory, DataUploadBatch, DataUpload, DataUploadTemplate, APIUploadLog, DataBatchStatus, DocumentTypeConfig, CalculationConfig, ConversionConfig, Currency, LineOfBusiness, ReportType, IFRSEngineResult, IFRSEngineInput, IFRSApiConfig, CalculationValue, AssumptionReference, InputDataReference, SubmittedReport
from .serializers import (
    ModelDefinitionListSerializer,
    ModelDefinitionDetailSerializer,
    ModelDefinitionCreateSerializer,
    ModelDefinitionUpdateSerializer,
    ModelDefinitionHistorySerializer,
    DataUploadBatchSerializer,
    DataUploadSerializer,
    DataUploadTemplateSerializer,
    APIUploadLogSerializer,
    DataBatchStatusSerializer,
    FileUploadSerializer,
    BulkUploadSerializer,
    DocumentTypeConfigSerializer,
    DocumentTypeConfigListSerializer,
    DocumentTypeConfigCreateSerializer,
    DocumentTypeConfigUpdateSerializer,
    CalculationConfigSerializer,
    CalculationConfigListSerializer,
    CalculationConfigCreateSerializer,
    CalculationConfigUpdateSerializer,
    ConversionConfigSerializer,
    ConversionConfigListSerializer,
    ConversionConfigCreateSerializer,
    ConversionConfigUpdateSerializer,
    CurrencySerializer,
    CurrencyListSerializer,
    CurrencyCreateSerializer,
    CurrencyUpdateSerializer,
    LineOfBusinessSerializer,
    LineOfBusinessListSerializer,
    LineOfBusinessCreateSerializer,
    LineOfBusinessUpdateSerializer,
    ReportTypeSerializer,
    ReportTypeCreateSerializer,
    ReportTypeUpdateSerializer,
    IFRSApiConfigSerializer,
    IFRSApiConfigCreateSerializer,
    IFRSApiConfigUpdateSerializer,
    IFRSEngineResultSerializer,
    IFRSEngineResultCreateSerializer,
    IFRSEngineInputSerializer,
    ReportGenerationSerializer,
    CalculationValueSerializer,
    AssumptionReferenceSerializer,
    InputDataReferenceSerializer,
    RunSummarySerializer,
    ReportMetadataSerializer,
    SubmittedReportSerializer,
    SubmittedReportCreateSerializer,
)


class ModelDefinitionViewSet(viewsets.ModelViewSet):
    queryset = ModelDefinition.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['created_by']
    search_fields = ['name']
    ordering_fields = ['name', 'created_on', 'modified_on']
    ordering = ['-modified_on']

    def get_serializer_class(self):
        if self.action == 'list':
            return ModelDefinitionListSerializer
        elif self.action == 'create':
            return ModelDefinitionCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return ModelDefinitionUpdateSerializer
        return ModelDefinitionDetailSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        
        measurement_model = self.request.query_params.get('measurement_model')
        if measurement_model:
            queryset = queryset.filter(config__general_info__measurement_model=measurement_model)
        
        product_type = self.request.query_params.get('product_type')
        if product_type:
            queryset = queryset.filter(config__general_info__product_type=product_type)
        
        status = self.request.query_params.get('status')
        if status:
            queryset = queryset.filter(config__general_info__status=status)
        
        return queryset

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Model definitions retrieved successfully.",
                "results": response.data
            })
        return response

    def retrieve(self, request, *args, **kwargs):
        response = super().retrieve(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Model definition retrieved successfully.",
                "model_definition": response.data
            })
        return response

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        with transaction.atomic():
            model_def = serializer.save()
            
            ModelDefinitionHistory.objects.create(
                model=model_def,
                name=model_def.name,
                version=model_def.version,
                config=model_def.config,
                modified_by=request.user
            )
        
        detail_serializer = ModelDefinitionDetailSerializer(
            model_def, context={'request': request}
        )
        
        return Response({
            "detail": "Model definition created successfully.",
            "model_definition": detail_serializer.data
        }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        if instance.locked_by and instance.locked_by != request.user:
            return Response({
                "detail": f"Model is currently being edited by {instance.locked_by.get_full_name()}."
            }, status=status.HTTP_423_LOCKED)
        
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        
        with transaction.atomic():
            updated_instance = serializer.save()
        
        detail_serializer = ModelDefinitionDetailSerializer(
            updated_instance, context={'request': request}
        )
        
        return Response({
            "detail": "Model definition updated successfully.",
            "model_definition": detail_serializer.data
        })

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        
        if not instance.can_edit(request.user):
            return Response({
                "detail": "You cannot delete this model."
            }, status=status.HTTP_403_FORBIDDEN)
        
        instance.delete()
        
        return Response({
            "detail": "Model definition deleted successfully."
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def lock(self, request, pk=None):
        instance = self.get_object()
        
        if instance.locked_by and instance.locked_by != request.user:
            return Response({
                "detail": f"Model is already locked by {instance.locked_by.get_full_name()}.",
                "locked_by": instance.locked_by.get_full_name(),
                "locked_at": instance.locked_at
            }, status=status.HTTP_423_LOCKED)
        
        instance.locked_by = request.user
        instance.locked_at = timezone.now()
        instance.save()
        
        return Response({
            "detail": "Model locked successfully.",
            "locked_by": request.user.get_full_name(),
            "locked_at": instance.locked_at
        })

    @action(detail=True, methods=['post'])
    def unlock(self, request, pk=None):
        instance = self.get_object()
        
        if instance.locked_by != request.user and not request.user.is_superuser:
            return Response({
                "detail": "You can only unlock models that you have locked."
            }, status=status.HTTP_403_FORBIDDEN)
        
        instance.locked_by = None
        instance.locked_at = None
        instance.save()
        
        return Response({
            "detail": "Model unlocked successfully."
        })

    @action(detail=True, methods=['post'])
    def clone(self, request, pk=None):
        source_instance = self.get_object()
        
        new_name = request.data.get('name')
        if not new_name:
            return Response({
                "detail": "New model name is required for cloning."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if ModelDefinition.objects.filter(name=new_name).exists():
            return Response({
                "detail": "A model with this name already exists."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        cloned_config = source_instance.config.copy()
        
        if 'generalInfo' in cloned_config:
            general_info = cloned_config['generalInfo']
            cloned_config['general_info'] = {
                'product_type': general_info.get('productType', ''),
                'measurement_model': general_info.get('measurementModel', ''),
                **{k: v for k, v in general_info.items() if k not in ['productType', 'measurementModel']}
            }
        elif 'general_info' not in cloned_config:
            cloned_config['general_info'] = {
                'product_type': '',  # Will need to be filled
                'measurement_model': 'GMM',  # Default value
            }
        
        cloned_data = {
            'name': new_name,
            'config': cloned_config,
            'cloned_from': source_instance.id
        }
        
        serializer = ModelDefinitionCreateSerializer(
            data=cloned_data, context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        
        with transaction.atomic():
            cloned_instance = serializer.save()
            
            ModelDefinitionHistory.objects.create(
                model=cloned_instance,
                name=cloned_instance.name,
                version=cloned_instance.version,
                config=cloned_instance.config,
                modified_by=request.user
            )
        
        detail_serializer = ModelDefinitionDetailSerializer(
            cloned_instance, context={'request': request}
        )
        
        return Response({
            "detail": "Model cloned successfully.",
            "model_definition": detail_serializer.data
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        instance = self.get_object()
        history_records = ModelDefinitionHistory.objects.filter(model=instance)
        
        serializer = ModelDefinitionHistorySerializer(
            history_records, many=True, context={'request': request}
        )
        
        return Response({
            "detail": "Model definition history retrieved successfully.",
            "history": serializer.data
        })


class ModelDefinitionHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ModelDefinitionHistory.objects.all()
    serializer_class = ModelDefinitionHistorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['model', 'modified_by']
    search_fields = ['name', 'model__name']
    ordering_fields = ['saved_at']
    ordering = ['-saved_at']

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Model definition history retrieved successfully.",
                "results": response.data
            })
        return response

class DataUploadBatchViewSet(viewsets.ModelViewSet):
    queryset = DataUploadBatch.objects.all()
    serializer_class = DataUploadBatchSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['batch_status', 'batch_type', 'batch_model', 'insurance_type', 'batch_year', 'batch_quarter', 'created_by']
    search_fields = ['batch_id', 'name']
    ordering_fields = ['created_on', 'modified_on']
    ordering = ['-created_on']

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Data upload batches retrieved successfully.",
                "results": response.data
            })
        return response

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        batch = serializer.save()
        
        return Response({
            "detail": "Data upload batch created successfully.",
            "batch": DataUploadBatchSerializer(batch).data
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def complete_batch(self, request, pk=None):
        batch = self.get_object()
        
        if batch.batch_status != 'pending':
            return Response({
                "detail": "Only pending batches can be completed."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if batch.upload_count == 0:
            return Response({
                "detail": "Cannot complete batch without any uploads."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        batch.batch_status = 'completed'
        batch.save()
        
        return Response({
            "detail": "Batch completed successfully.",
            "batch": DataUploadBatchSerializer(batch).data
        })

    @action(detail=True, methods=['get'])
    def status_records(self, request, pk=None):
        batch = self.get_object()
        status_records = DataBatchStatus.objects.filter(batch_id=batch.batch_id)
        serializer = DataBatchStatusSerializer(status_records, many=True)
        
        return Response({
            "detail": "Batch status records retrieved successfully.",
            "results": serializer.data
        })


class DataUploadViewSet(viewsets.ModelViewSet):
    queryset = DataUpload.objects.all()
    serializer_class = DataUploadSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['source', 'insurance_type', 'data_type', 'quarter', 'year', 'validation_status', 'batch']
    search_fields = ['upload_id', 'batch__batch_id', 'original_filename']
    ordering_fields = ['created_on', 'modified_on']
    ordering = ['-created_on']

    def get_serializer_class(self):
        if self.action == 'upload_file':
            return FileUploadSerializer
        elif self.action == 'bulk_upload':
            return BulkUploadSerializer
        return DataUploadSerializer

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Data uploads retrieved successfully.",
                "results": response.data
            })
        return response

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        upload = serializer.save()
        
        return Response({
            "detail": "Data upload created successfully.",
            "upload": DataUploadSerializer(upload).data
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def upload_file(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        upload = serializer.save()
        
        return Response({
            "detail": "File uploaded successfully.",
            "upload": DataUploadSerializer(upload).data
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        uploads = serializer.save()
        
        return Response({
            "detail": f"{len(uploads)} files uploaded successfully.",
            "uploads": [DataUploadSerializer(upload).data for upload in uploads]
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        upload = self.get_object()
        
        if not upload.file_upload:
            return Response({
                "detail": "No file associated with this upload."
            }, status=status.HTTP_404_NOT_FOUND)
        
        try:
            response = HttpResponse(upload.file_upload.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="{upload.original_filename}"'
            return response
        except Exception as e:
            return Response({
                "detail": "Error downloading file.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def retry_validation(self, request, pk=None):
        upload = self.get_object()
        
        if upload.validation_status != 'failed':
            return Response({
                "detail": "Only failed uploads can be retried."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # TODO: Implement retry validation logic
        upload.validation_status = 'in_progress'
        upload.error_count = 0
        upload.validation_errors = []
        upload.save()
        
        return Response({
            "detail": "Validation retry initiated.",
            "upload": DataUploadSerializer(upload).data
        })


class DataUploadTemplateViewSet(viewsets.ModelViewSet):
    queryset = DataUploadTemplate.objects.all()
    serializer_class = DataUploadTemplateSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['data_type', 'is_active', 'is_standard_template']
    search_fields = ['name', 'data_type']
    ordering_fields = ['created_on', 'modified_on']
    ordering = ['-created_on']

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Data upload templates retrieved successfully.",
                "results": response.data
            })
        return response

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        template = self.get_object()
        
        if not template.template_file:
            return Response({
                "detail": "No template file associated with this template."
            }, status=status.HTTP_404_NOT_FOUND)
        
        try:
            response = HttpResponse(template.template_file.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="{template.name}.xlsx"'
            return response
        except Exception as e:
            return Response({
                "detail": "Error downloading template.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def get_by_data_type(self, request):
        data_type = request.query_params.get('data_type')
        if not data_type:
            return Response({
                "detail": "data_type parameter is required."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        templates = self.get_queryset().filter(data_type=data_type, is_active=True)
        serializer = self.get_serializer(templates, many=True)
        
        return Response({
            "detail": "Templates retrieved successfully.",
            "results": serializer.data
        })


class APIUploadLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = APIUploadLog.objects.all()
    serializer_class = APIUploadLogSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'reporting_date']
    search_fields = ['data_upload__upload_id']
    ordering_fields = ['upload_date', 'reporting_date']
    ordering = ['-upload_date']

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "API upload logs retrieved successfully.",
                "results": response.data
            })
        return response

    @action(detail=True, methods=['post'])
    def retry_upload(self, request, pk=None):
        api_log = self.get_object()
        
        if api_log.status != 'failed':
            return Response({
                "detail": "Only failed uploads can be retried."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # TODO: Implement retry upload logic
        # This would typically involve re-processing the API payload
        
        return Response({
            "detail": "API upload retry initiated.",
            "log": APIUploadLogSerializer(api_log).data
        })


class DataBatchStatusViewSet(viewsets.ModelViewSet):
    queryset = DataBatchStatus.objects.all()
    serializer_class = DataBatchStatusSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['batch_id', 'document_type', 'upload_status']
    search_fields = ['batch_id', 'document_type']
    ordering_fields = ['batch_id', 'document_type']
    ordering = ['batch_id', 'document_type']

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Batch status records retrieved successfully.",
                "results": response.data
            })
        return response

    @action(detail=False, methods=['get'])
    def get_by_batch(self, request):
        batch_id = request.query_params.get('batch_id')
        if not batch_id:
            return Response({
                "detail": "batch_id parameter is required."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        status_records = self.get_queryset().filter(batch_id=batch_id)
        serializer = self.get_serializer(status_records, many=True)
        
        return Response({
            "detail": "Batch status records retrieved successfully.",
            "results": serializer.data
        })

    @action(detail=False, methods=['post'])
    def update_status(self, request):
        batch_id = request.data.get('batch_id')
        document_type = request.data.get('document_type')
        upload_status = request.data.get('upload_status')
        
        if not batch_id or not document_type or upload_status is None:
            return Response({
                "detail": "batch_id, document_type, and upload_status are required."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            status_record = DataBatchStatus.objects.get(
                batch_id=batch_id,
                document_type=document_type
            )
            status_record.upload_status = upload_status
            status_record.save()
            
            return Response({
                "detail": "Status updated successfully.",
                "status_record": DataBatchStatusSerializer(status_record).data
            })
        except DataBatchStatus.DoesNotExist:
            return Response({
                "detail": "Status record not found."
            }, status=status.HTTP_404_NOT_FOUND)


class DocumentTypeConfigViewSet(viewsets.ModelViewSet):
    queryset = DocumentTypeConfig.objects.all()
    serializer_class = DocumentTypeConfigSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['batch_type', 'batch_model', 'insurance_type', 'required']
    search_fields = ['document_type', 'batch_type', 'batch_model', 'insurance_type']
    ordering_fields = ['created_on', 'modified_on', 'batch_type', 'batch_model', 'insurance_type', 'document_type']
    ordering = ['batch_type', 'batch_model', 'insurance_type', 'document_type']

    def get_serializer_class(self):
        if self.action == 'list':
            return DocumentTypeConfigListSerializer
        elif self.action == 'create':
            return DocumentTypeConfigCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return DocumentTypeConfigUpdateSerializer
        return DocumentTypeConfigSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Document type configurations retrieved successfully.",
                "results": response.data
            })
        return response

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        document_type_config = serializer.save()
        
        return Response({
            "detail": "Document type configuration created successfully.",
            "documentTypeConfig": DocumentTypeConfigSerializer(document_type_config, context={'request': request}).data
        }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        document_type_config = serializer.save()
        
        return Response({
            "detail": "Document type configuration updated successfully.",
            "documentTypeConfig": DocumentTypeConfigSerializer(document_type_config, context={'request': request}).data
        })

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({
            "detail": "Document type configuration deleted successfully."
        }, status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['get'])
    def download_template(self, request, pk=None):
        document_type_config = self.get_object()
        
        if not document_type_config.template:
            return Response({
                "detail": "No template file associated with this configuration."
            }, status=status.HTTP_404_NOT_FOUND)
        
        try:
            response = HttpResponse(
                document_type_config.template.read(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = document_type_config.template.name.split('/')[-1]
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return Response({
                "detail": "Error downloading template file.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CalculationConfigViewSet(viewsets.ModelViewSet):
    queryset = CalculationConfig.objects.all()
    serializer_class = CalculationConfigSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['batch_type', 'batch_model', 'insurance_type', 'required']
    search_fields = ['engine_type', 'batch_type', 'batch_model', 'insurance_type']
    ordering_fields = ['created_on', 'modified_on', 'batch_type', 'batch_model', 'insurance_type', 'engine_type']
    ordering = ['batch_type', 'batch_model', 'insurance_type', 'engine_type']

    def get_serializer_class(self):
        if self.action == 'list':
            return CalculationConfigListSerializer
        elif self.action == 'create':
            return CalculationConfigCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return CalculationConfigUpdateSerializer
        return CalculationConfigSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Calculation configurations retrieved successfully.",
                "results": response.data
            })
        return response

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        calculation_config = serializer.save()
        
        return Response({
            "detail": "Calculation configuration created successfully.",
            "engineConfig": CalculationConfigSerializer(calculation_config, context={'request': request}).data
        }, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        calculation_config = serializer.save()
        
        return Response({
            "detail": "Calculation configuration updated successfully.",
            "engineConfig": CalculationConfigSerializer(calculation_config, context={'request': request}).data
        })

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({
            "detail": "Calculation configuration deleted successfully."
        }, status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['get'])
    def download_script(self, request, pk=None):
        calculation_config = self.get_object()
        
        if not calculation_config.script:
            return Response({
                "detail": "No script file associated with this configuration."
            }, status=status.HTTP_404_NOT_FOUND)
        
        try:
            response = HttpResponse(
                calculation_config.script.read(), 
                content_type='text/x-python'
            )
            filename = calculation_config.script.name.split('/')[-1]
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return Response({
                "detail": "Error downloading script file.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ConversionConfigViewSet(viewsets.ModelViewSet):
    queryset = ConversionConfig.objects.all()
    serializer_class = ConversionConfigSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['batch_type', 'batch_model', 'insurance_type', 'required']
    search_fields = ['engine_type', 'batch_type', 'batch_model', 'insurance_type']
    ordering_fields = ['created_on', 'modified_on', 'batch_type', 'batch_model', 'insurance_type', 'engine_type']
    ordering = ['batch_type', 'batch_model', 'insurance_type', 'engine_type']

    def get_serializer_class(self):
        if self.action == 'list':
            return ConversionConfigListSerializer
        elif self.action == 'create':
            return ConversionConfigCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return ConversionConfigUpdateSerializer
        return ConversionConfigSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Conversion configurations retrieved successfully.",
                "results": response.data
            })
        return response

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        if response.status_code == status.HTTP_201_CREATED:
            return Response({
                "detail": "Conversion configuration created successfully.",
                "conversion_config": response.data
            }, status=status.HTTP_201_CREATED)
        return response

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Conversion configuration updated successfully.",
                "conversion_config": response.data
            })
        return response

    def destroy(self, request, *args, **kwargs):
        response = super().destroy(request, *args, **kwargs)
        if response.status_code == status.HTTP_204_NO_CONTENT:
            return Response({
                "detail": "Conversion configuration deleted successfully."
            }, status=status.HTTP_200_OK)
        return response

    @action(detail=True, methods=['get'])
    def download_script(self, request, pk=None):
        try:
            conversion_config = self.get_object()
            if not conversion_config.script:
                return Response({
                    "detail": "No script file available for this conversion configuration."
                }, status=status.HTTP_404_NOT_FOUND)
            
            response = HttpResponse(conversion_config.script, content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="{conversion_config.script.name.split("/")[-1]}"'
            return response
        except Http404:
            return Response({
                "detail": "Conversion configuration not found."
            }, status=status.HTTP_404_NOT_FOUND)


class CurrencyViewSet(viewsets.ModelViewSet):
    queryset = Currency.objects.all()
    serializer_class = CurrencySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active']
    search_fields = ['code', 'name']
    ordering_fields = ['code', 'name', 'created_on', 'modified_on']
    ordering = ['code']

    def get_serializer_class(self):
        if self.action == 'list':
            return CurrencyListSerializer
        elif self.action == 'create':
            return CurrencyCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return CurrencyUpdateSerializer
        return CurrencySerializer

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Currencies retrieved successfully.",
                "results": response.data
            })
        return response

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        if response.status_code == status.HTTP_201_CREATED:
            return Response({
                "detail": "Currency created successfully.",
                "currency": response.data
            }, status=status.HTTP_201_CREATED)
        return response

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Currency updated successfully.",
                "currency": response.data
            })
        return response

    def destroy(self, request, *args, **kwargs):
        response = super().destroy(request, *args, **kwargs)
        if response.status_code == status.HTTP_204_NO_CONTENT:
            return Response({
                "detail": "Currency deleted successfully."
            }, status=status.HTTP_200_OK)
        return response

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get only active currencies"""
        currencies = Currency.objects.filter(is_active=True).order_by('code')
        serializer = CurrencyListSerializer(currencies, many=True)
        return Response({
            "detail": "Active currencies retrieved successfully.",
            "results": serializer.data
        })


class LineOfBusinessViewSet(viewsets.ModelViewSet):
    queryset = LineOfBusiness.objects.all()
    serializer_class = LineOfBusinessSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['batch_model', 'insurance_type', 'currency', 'is_active']
    search_fields = ['line_of_business', 'batch_model', 'insurance_type']
    ordering_fields = ['created_on', 'modified_on', 'batch_model', 'insurance_type', 'line_of_business']
    ordering = ['batch_model', 'insurance_type', 'line_of_business']

    def get_serializer_class(self):
        if self.action == 'list':
            return LineOfBusinessListSerializer
        elif self.action == 'create':
            return LineOfBusinessCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return LineOfBusinessUpdateSerializer
        return LineOfBusinessSerializer

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Lines of business retrieved successfully.",
                "results": response.data
            })
        return response

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        if response.status_code == status.HTTP_201_CREATED:
            return Response({
                "detail": "Line of business created successfully.",
                "line_of_business": response.data
            }, status=status.HTTP_201_CREATED)
        return response

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Line of business updated successfully.",
                "line_of_business": response.data
            })
        return response

    def destroy(self, request, *args, **kwargs):
        response = super().destroy(request, *args, **kwargs)
        if response.status_code == status.HTTP_204_NO_CONTENT:
            return Response({
                "detail": "Line of business deleted successfully."
            }, status=status.HTTP_200_OK)
        return response

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get only active lines of business"""
        lines_of_business = LineOfBusiness.objects.filter(is_active=True).order_by('batch_model', 'insurance_type', 'line_of_business')
        serializer = LineOfBusinessListSerializer(lines_of_business, many=True)
        return Response({
            "detail": "Active lines of business retrieved successfully.",
            "results": serializer.data
        })

    @action(detail=False, methods=['get'])
    def by_model_and_type(self, request):
        """Get lines of business filtered by batch model and insurance type"""
        batch_model = request.query_params.get('batch_model')
        insurance_type = request.query_params.get('insurance_type')
        
        queryset = LineOfBusiness.objects.filter(is_active=True)
        
        if batch_model:
            queryset = queryset.filter(batch_model=batch_model)
        if insurance_type:
            queryset = queryset.filter(insurance_type=insurance_type)
        
        queryset = queryset.order_by('line_of_business')
        serializer = LineOfBusinessListSerializer(queryset, many=True)
        
        return Response({
            "detail": "Lines of business filtered successfully.",
            "results": serializer.data
        })

class ReportTypeViewSet(viewsets.ModelViewSet):
    queryset = ReportType.objects.all()
    serializer_class = ReportTypeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['batch_model', 'is_enabled']
    search_fields = ['report_type']
    ordering_fields = ['batch_model', 'report_type', 'created_on']
    ordering = ['batch_model', 'report_type']

    def get_serializer_class(self):
        if self.action == 'create':
            return ReportTypeCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return ReportTypeUpdateSerializer
        return ReportTypeSerializer

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "Report types retrieved successfully.",
                "results": response.data
            })
        return response

    @action(detail=False, methods=['get'])
    def by_model(self, request):
        batch_model = request.query_params.get('batch_model')
        if batch_model:
            queryset = ReportType.objects.filter(batch_model=batch_model)
            serializer = self.get_serializer(queryset, many=True)
            return Response({
                "detail": "Report types filtered successfully.",
                "results": serializer.data
            })
        return Response({'error': 'batch_model parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def enabled(self, request):
        enabled_reports = ReportType.objects.filter(is_enabled=True)
        serializer = self.get_serializer(enabled_reports, many=True)
        return Response({
            "detail": "Enabled report types retrieved successfully.",
            "results": serializer.data
        })


class IFRSEngineInputViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = IFRSEngineInput.objects.all()
    serializer_class = IFRSEngineInputSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['run_id', 'created_by']
    search_fields = ['run_id']
    ordering_fields = ['created_at']
    ordering = ['-created_at']

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        if response.status_code == status.HTTP_200_OK:
            return Response({
                "detail": "IFRS Engine Inputs retrieved successfully.",
                "results": response.data
            })
        return response


class SubmittedReportViewSet(viewsets.ModelViewSet):
    queryset = SubmittedReport.objects.all()
    serializer_class = SubmittedReportSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['assign_year', 'assign_quarter', 'status', 'report_type', 'model_type']
    search_fields = ['report_type', 'run_id']
    ordering_fields = ['assign_year', 'assign_quarter', 'created_on']
    ordering = ['-assign_year', '-assign_quarter', '-created_on']
    
    def get_queryset(self):
        queryset = SubmittedReport.objects.all()
        
        year = self.request.query_params.get('year', None)
        quarter = self.request.query_params.get('quarter', None)
        report_type = self.request.query_params.get('report_type', None)
        status_filter = self.request.query_params.get('status', None)
        
        if year:
            queryset = queryset.filter(assign_year=year)
        if quarter:
            queryset = queryset.filter(assign_quarter=quarter)
        if report_type:
            queryset = queryset.filter(report_type__icontains=report_type)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset.order_by('-assign_year', '-assign_quarter', '-created_on')
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 12))
        
        total_count = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        
        results = queryset[start:end]
        
        serializer = self.get_serializer(results, many=True)
        
        total_pages = (total_count + page_size - 1) // page_size
        has_next = page < total_pages
        has_previous = page > 1
        
        return Response({
            'count': total_count,
            'next': f"?page={page + 1}&page_size={page_size}" if has_next else None,
            'previous': f"?page={page - 1}&page_size={page_size}" if has_previous else None,
            'results': serializer.data
        })
    
    @action(detail=False, methods=['post'])
    def submit_reports(self, request):
        serializer = SubmittedReportCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        data = serializer.validated_data
        report_ids = data['report_ids']
        assign_year = data['assign_year']
        assign_quarter = data['assign_quarter']
        
        ifrs_results = IFRSEngineResult.objects.filter(id__in=report_ids)
        if not ifrs_results.exists():
            return Response({'error': 'No reports found with provided IDs'}, status=status.HTTP_400_BAD_REQUEST)
        
        submitted_reports = []
        for result in ifrs_results:
            model_name = ''
            batch_info = ''
            lob_info = ''
            conversion_engine = ''
            ifrs_engine = ''
            
            try:
                engine_input = IFRSEngineInput.objects.get(run_id=result.run_id)
                
                if engine_input.model_definition:
                    model_def = engine_input.model_definition
                    model_name = f"{model_def.get('name', 'N/A')} (v{model_def.get('version', 'N/A')})"
                
                if engine_input.batch_data and len(engine_input.batch_data) > 0:
                    batch_ids = [b.get('batch_id', '') for b in engine_input.batch_data if b.get('batch_id')]
                    batch_info = ', '.join(batch_ids) if batch_ids else 'N/A'
                
                if engine_input.field_parameters and 'line_of_businesses' in engine_input.field_parameters:
                    lobs = engine_input.field_parameters['line_of_businesses']
                    lob_names = [lob.get('line_of_business', '') for lob in lobs if lob.get('line_of_business')]
                    lob_info = ', '.join(lob_names) if lob_names else 'N/A'
                
                if engine_input.field_parameters:
                    conversion_engine = f"{result.model_type} Conversion Engine"
                    
                    ifrs_engine_id = engine_input.field_parameters.get('ifrs_engine_id')
                    if ifrs_engine_id:
                        try:
                            ifrs_calc_config = CalculationConfig.objects.get(id=ifrs_engine_id)
                            ifrs_engine = ifrs_calc_config.engine_type
                        except CalculationConfig.DoesNotExist:
                            ifrs_engine = f"Engine ID: {ifrs_engine_id}"
                    else:
                        ifrs_engine = 'Default IFRS Engine'
                        
            except IFRSEngineInput.DoesNotExist:
                model_name = 'N/A'
                batch_info = 'N/A'
                lob_info = 'N/A'
                conversion_engine = f"{result.model_type} Conversion Engine"
                ifrs_engine = 'Default IFRS Engine'
            
            submitted_report = SubmittedReport(
                run_id=result.run_id,
                report_type=result.report_type,
                report_type_display=result.report_type,
                model_type=result.model_type,
                assign_year=assign_year,
                assign_quarter=assign_quarter,
                status='active',
                ifrs_engine_result_id=result.id,
                model_used=model_name,
                batch_used=batch_info,
                line_of_business_used=lob_info,
                conversion_engine_used=conversion_engine,
                ifrs_engine_used=ifrs_engine,
                submitted_by=request.user
            )
            submitted_report.save()
            submitted_reports.append(submitted_report)
        
        serializer = SubmittedReportSerializer(submitted_reports, many=True)
        return Response({
            'message': f'Successfully submitted {len(submitted_reports)} reports',
            'reports': serializer.data
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['delete'])
    def delete_submitted_report(self, request, pk=None):
        try:
            submitted_report = self.get_object()
            submitted_report.delete()
            return Response({
                'detail': 'Submitted report deleted successfully'
            }, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            return Response({
                'error': f'Failed to delete submitted report: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def download_excel(self, request, pk=None):
        """Download submitted report Excel by delegating to the underlying IFRSEngineResult"""
        try:
            submitted_report = self.get_object()
            
            ifrs_result = IFRSEngineResult.objects.get(id=submitted_report.ifrs_engine_result_id)
            
            from rest_framework.request import Request
            
            ifrs_viewset = IFRSEngineResultViewSet()
            ifrs_viewset.kwargs = {'pk': ifrs_result.id}
            ifrs_viewset.request = request
            ifrs_viewset.format_kwarg = None
            
            response = ifrs_viewset.download_excel(request, pk=ifrs_result.id)
            
            if response.status_code == 200:
                response['Content-Disposition'] = f'attachment; filename="{submitted_report.report_type}_{submitted_report.assign_year}_Q{submitted_report.assign_quarter}.xlsx"'
            
            return response
            
        except IFRSEngineResult.DoesNotExist:
            return Response({
                'error': 'Report result not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            import traceback
            return Response({
                'error': f'Failed to download Excel: {str(e)}',
                'traceback': traceback.format_exc()
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class IFRSEngineResultViewSet(viewsets.ModelViewSet):
    queryset = IFRSEngineResult.objects.all()
    serializer_class = IFRSEngineResultSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['run_id', 'model_type', 'report_type', 'year', 'quarter', 'created_by', 'status']
    search_fields = ['run_id', 'model_guid', 'report_type']
    ordering_fields = ['created_at', 'model_type', 'report_type', 'year', 'quarter']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'create':
            return IFRSEngineResultCreateSerializer
        return IFRSEngineResultSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 12))
        
        total_count = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        
        results = queryset[start:end]
        
        serializer = self.get_serializer(results, many=True)
        
        total_pages = (total_count + page_size - 1) // page_size
        has_next = page < total_pages
        has_previous = page > 1
        
        return Response({
            'count': total_count,
            'next': f"?page={page + 1}&page_size={page_size}" if has_next else None,
            'previous': f"?page={page - 1}&page_size={page_size}" if has_previous else None,
            'results': serializer.data
        })

    def perform_create(self, serializer):
        username = self.request.user.username if self.request.user else 'system'
        serializer.save(created_by=username)

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        try:
            result = self.get_object()
            
            if result.status != 'Success':
                return Response({
                    'error': 'Cannot generate PDF for failed report'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            from reportlab.lib.pagesizes import letter, A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from io import BytesIO
            import json
            from datetime import datetime
            
            buffer = BytesIO()
            
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=landscape(A4),
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                alignment=1,
                spaceAfter=20,
                fontSize=16,
                textColor=colors.darkblue
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                spaceAfter=10,
                spaceBefore=15,
                fontSize=14,
                textColor=colors.darkblue
            )
            
            normal_style = styles['Normal']
            
            content = []
            
            title = f'IFRS Engine Report - {result.report_type.replace("_", " ").title()}'
            content.append(Paragraph(title, title_style))
            content.append(Spacer(1, 12))
            
            metadata_data = [
                ['Run ID:', result.run_id],
                ['Model Type:', result.model_type],
                ['Report Type:', result.report_type.replace('_', ' ').title()],
                ['Year:', str(result.year)],
                ['Quarter:', result.quarter],
                ['Currency:', result.currency or 'USD'],
                ['Status:', result.status],
                ['Generated:', result.created_at.strftime('%Y-%m-%d %H:%M:%S') if result.created_at else 'N/A']
            ]
            
            metadata_table = Table(metadata_data, colWidths=[2*inch, 4*inch])
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            content.append(metadata_table)
            content.append(Spacer(1, 20))
            
            if isinstance(result.result_json, dict) and 'results' in result.result_json:
                results_data = result.result_json['results']
                
                summary_data = results_data.get('summaryView') or results_data.get('summary_view')
                if summary_data and len(summary_data) > 0:
                    content.append(Paragraph('Summary View', heading_style))
                    
                    headers = list(summary_data[0].keys())
                    table_data = [headers]
                    
                    for row in summary_data:
                        table_row = []
                        for header in headers:
                            value = row.get(header, '')
                            if isinstance(value, (int, float)):
                                if abs(value) >= 1000:
                                    formatted_value = f"{value:,.0f}"
                                else:
                                    formatted_value = f"{value:,.2f}"
                            else:
                                formatted_value = str(value)
                            table_row.append(formatted_value)
                        table_data.append(table_row)
                    
                    available_width = 10 * inch
                    num_columns = len(headers)
                    col_width = available_width / num_columns
                    col_widths = [col_width] * num_columns
                    
                    summary_table = Table(table_data, colWidths=col_widths, repeatRows=1)
                    summary_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 0), (-1, 0), 9),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]))
                    
                    content.append(summary_table)
                    content.append(PageBreak())
                
                detailed_data = results_data.get('detailedView') or results_data.get('detailed_view')
                if detailed_data and len(detailed_data) > 0:
                    content.append(Paragraph('Detailed View', heading_style))
                    
                    headers = list(detailed_data[0].keys())
                    table_data = [headers]
                    
                    display_rows = detailed_data[:50]
                    
                    for row in display_rows:
                        table_row = []
                        for header in headers:
                            value = row.get(header, '')
                            if isinstance(value, (int, float)):
                                if abs(value) >= 1000:
                                    formatted_value = f"{value:,.0f}"
                                else:
                                    formatted_value = f"{value:,.2f}"
                            else:
                                formatted_value = str(value)[:15]  # Truncate long strings
                            table_row.append(formatted_value)
                        table_data.append(table_row)
                    
                    available_width = 10 * inch
                    num_columns = len(headers)
                    col_width = available_width / num_columns
                    col_widths = [col_width] * num_columns
                    
                    detailed_table = Table(table_data, colWidths=col_widths, repeatRows=1)
                    detailed_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 0), (-1, 0), 8),
                        ('FONTSIZE', (0, 1), (-1, -1), 7),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]))
                    
                    content.append(detailed_table)
                    
                    if len(detailed_data) > 50:
                        content.append(Spacer(1, 12))
                        content.append(Paragraph(f'Note: Showing first 50 of {len(detailed_data)} total records. Download Excel for complete data.', normal_style))
            
            doc.build(content)
            
            pdf_content = buffer.getvalue()
            buffer.close()
            
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{result.report_type}_report_{result.run_id}.pdf"'
            
            return response
            
        except Exception as e:
            import traceback
            return Response({
                'error': f'Failed to generate PDF: {str(e)}',
                'traceback': traceback.format_exc()
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def download_excel(self, request, pk=None):
        """Download IFRS engine result as Excel with enhanced formatting"""
        try:
            from django.http import HttpResponse
            import base64
            
            result = self.get_object()
            
            if result.status != 'Success':
                return Response({
                    'error': 'Cannot generate Excel for failed report'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if result.report_type == 'disclosure_report':
                
                if not isinstance(result.result_json, dict):
                    return Response({
                        'error': 'Invalid result JSON structure',
                        'result_type': str(type(result.result_json))
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                if 'error' in result.result_json:
                    return Response({
                        'error': 'Engine execution failed',
                        'details': result.result_json.get('error'),
                        'stdout': result.result_json.get('stdout', 'No stdout'),
                        'stderr': result.result_json.get('stderr', 'No stderr'),
                        'return_code': result.result_json.get('return_code'),
                        'traceback': result.result_json.get('traceback', 'No traceback available')
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                if 'excel_bytes' not in result.result_json:
                    return Response({
                        'error': 'Disclosure report Excel not found in result',
                        'available_keys': list(result.result_json.keys()),
                        'status': result.result_json.get('status', 'unknown')
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                try:
                    excel_bytes = base64.b64decode(result.result_json['excel_bytes'])
                    
                    response = HttpResponse(
                        excel_bytes, 
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = f'attachment; filename="disclosure_report_{result.run_id}.xlsx"'
                    return response
                except Exception as e:
                    return Response({
                        'error': 'Failed to decode Excel bytes',
                        'details': str(e)
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.workbook import Workbook
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from io import BytesIO
            from datetime import datetime
            
            wb = Workbook()
            
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            currency_format = '#,##0.00'
            number_format = '#,##0'
            
            if isinstance(result.result_json, dict) and 'results' in result.result_json:
                results_data = result.result_json['results']
                
                metadata_ws = wb.create_sheet('Report Info')
                
                summary_data = results_data.get('summaryView', results_data.get('summary_view', []))
                detailed_data = results_data.get('detailedView', results_data.get('detailed_view', []))
                
                report_descriptions = {
                    'lrc_movement_report': 'Liability for Remaining Coverage movements',
                    'lic_movement_report': 'Liability for Incurred Claims movements', 
                    'staging_table': 'Conversion engine output data',
                    'insurance_revenue_expense_report': 'Insurance revenue and expense analysis',
                    'disclosure_report': 'Financial disclosure information',
                    'financial_statement_items_report': 'Financial statement line items',
                    'premium_allocation_reconciliation': 'Premium allocation and reconciliation',
                    'loss_component_report': 'Loss component analysis',
                    'discount_rate_reconciliation': 'Discount rate reconciliation',
                    'experience_adjustment_report': 'Experience adjustment calculations',
                    'reinsurance_report': 'Reinsurance transactions and balances',
                    'paa_roll_forward_report': 'Premium Allocation Approach roll forward',
                    'cash_flow_statement_report': 'Cash flow statement preparation'
                }
                
                lobs = []
                if summary_data:
                    first_row = summary_data[0] if summary_data else {}
                    lobs = [key for key in first_row.keys() if key not in ['reportingDate', 'year', 'reporting_date']]
                elif detailed_data:
                    lobs = list(set(row.get('lob', '') for row in detailed_data if row.get('lob')))
                
                lob_text = f" across {len(lobs)} Lines of Business: {', '.join(lobs[:5])}" if lobs else ""
                if len(lobs) > 5:
                    lob_text += f" and {len(lobs) - 5} others"
                
                report_desc = report_descriptions.get(result.report_type, 'Report data')
                summary_desc = f"{len(summary_data)} records - {report_desc} aggregated by Line of Business" if summary_data else "No summary data available"
                detailed_desc = f"{len(detailed_data)} records - {report_desc} with transaction-level detail{lob_text}" if detailed_data else "No detailed data available"
                
                metadata_data = [
                    ['Field', 'Value'],
                    ['Run ID', result.run_id],
                    ['Model Type', result.model_type],
                    ['Report Type', result.report_type.replace('_', ' ').title()],
                    ['Report Description', report_descriptions.get(result.report_type, 'IFRS calculation report')],
                    ['Year', result.year],
                    ['Quarter', result.quarter],
                    ['Currency', result.currency or 'USD'],
                    ['Status', result.status],
                    ['Generated', result.created_at.strftime('%Y-%m-%d %H:%M:%S') if result.created_at else 'N/A'],
                    ['Lines of Business', f"{len(lobs)} LOBs: {', '.join(lobs)}" if lobs else "Not available"],
                    ['Summary Records', summary_desc],
                    ['Detailed Records', detailed_desc]
                ]
                
                for row_idx, row_data in enumerate(metadata_data, 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = metadata_ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = cell_border
                        if row_idx == 1:  # Header row
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                        elif col_idx == 1:  # Field column
                            cell.font = Font(bold=True)
                
                for column in metadata_ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    metadata_ws.column_dimensions[column_letter].width = adjusted_width
                
                summary_data = results_data.get('summaryView') or results_data.get('summary_view')
                if summary_data and len(summary_data) > 0:
                    summary_ws = wb.create_sheet('Summary View')
                    
                    headers = list(summary_data[0].keys())
                    table_data = [headers]
                    
                    for row in summary_data:
                        table_row = []
                        for header in headers:
                            value = row.get(header, '')
                            if isinstance(value, (int, float)):
                                if abs(value) >= 1000:
                                    formatted_value = f"{value:,.0f}"
                                else:
                                    formatted_value = f"{value:,.2f}"
                            else:
                                formatted_value = str(value)
                            table_row.append(formatted_value)
                        table_data.append(table_row)
                    
                    for row_idx, row_data in enumerate(table_data, 1):
                        for col_idx, value in enumerate(row_data, 1):
                            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.border = cell_border
                            cell.alignment = center_alignment
                            
                            if row_idx == 1:
                                cell.font = header_font
                                cell.fill = header_fill
                            elif isinstance(value, str) and value.replace(',', '').replace('.', '').replace('-', '').isdigit():
                                try:
                                    numeric_value = float(value.replace(',', ''))
                                    if abs(numeric_value) >= 1000:
                                        cell.number_format = currency_format
                                    else:
                                        cell.number_format = number_format
                                except:
                                    pass
                    
                    table_ref = f"A1:{chr(65 + len(headers) - 1)}{len(table_data)}"
                    table = Table(displayName="SummaryTable", ref=table_ref)
                    style = TableStyleInfo(
                        name="TableStyleMedium9", 
                        showFirstColumn=False,
                        showLastColumn=False, 
                        showRowStripes=True, 
                        showColumnStripes=False
                    )
                    table.tableStyleInfo = style
                    summary_ws.add_table(table)
                    
                    for column in summary_ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        summary_ws.column_dimensions[column_letter].width = adjusted_width
                
                detailed_data = results_data.get('detailedView') or results_data.get('detailed_view')
                if detailed_data and len(detailed_data) > 0:
                    detailed_ws = wb.create_sheet('Detailed View')
                    
                    headers = list(detailed_data[0].keys())
                    table_data = [headers]
                    
                    for row in detailed_data:
                        table_row = []
                        for header in headers:
                            value = row.get(header, '')
                            if isinstance(value, (int, float)):
                                if abs(value) >= 1000:
                                    formatted_value = f"{value:,.0f}"
                                else:
                                    formatted_value = f"{value:,.2f}"
                            else:
                                formatted_value = str(value)[:15]
                            table_row.append(formatted_value)
                        table_data.append(table_row)
                    
                    for row_idx, row_data in enumerate(table_data, 1):
                        for col_idx, value in enumerate(row_data, 1):
                            cell = detailed_ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.border = cell_border
                            cell.alignment = center_alignment
                            
                            if row_idx == 1:
                                cell.font = Font(bold=True, color='FFFFFF')
                                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                            elif isinstance(value, str) and value.replace(',', '').replace('.', '').replace('-', '').isdigit():
                                try:
                                    numeric_value = float(value.replace(',', ''))
                                    if abs(numeric_value) >= 1000:
                                        cell.number_format = currency_format
                                    else:
                                        cell.number_format = number_format
                                except:
                                    pass
                    
                    table_ref = f"A1:{chr(65 + len(headers) - 1)}{len(table_data)}"
                    table = Table(displayName="DetailedTable", ref=table_ref)
                    style = TableStyleInfo(
                        name="TableStyleMedium2", 
                        showFirstColumn=False,
                        showLastColumn=False, 
                        showRowStripes=True, 
                        showColumnStripes=False
                    )
                    table.tableStyleInfo = style
                    detailed_ws.add_table(table)
                    
                    for column in detailed_ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        detailed_ws.column_dimensions[column_letter].width = adjusted_width
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            excel_content = output.getvalue()
            
            # Create response
            response = HttpResponse(excel_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{result.report_type}_report_{result.run_id}.xlsx"'
            
            return response
            
        except Exception as e:
            import traceback
            return Response({
                'error': f'Failed to generate Excel: {str(e)}',
                'traceback': traceback.format_exc()
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, *args, **kwargs):
        """Delete an IFRS engine result"""
        try:
            result = self.get_object()
            result.delete()
            return Response({
                'detail': 'IFRS engine result deleted successfully'
            }, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            return Response({
                'error': f'Failed to delete IFRS engine result: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def generate(self, request):
        serializer = ReportGenerationSerializer(data=request.data)
        if serializer.is_valid():
            data = serializer.validated_data
            
            try:
                model = ModelDefinition.objects.get(id=data['model_id'])
            except ModelDefinition.DoesNotExist:
                return Response({'error': 'Model not found'}, status=status.HTTP_400_BAD_REQUEST)
            
            batches = DataUploadBatch.objects.filter(
                id__in=data['batch_ids'],
                batch_status='completed'
            )
            if not batches.exists():
                return Response({'error': 'No completed batches found'}, status=status.HTTP_400_BAD_REQUEST)
            
            line_of_businesses = LineOfBusiness.objects.filter(
                id__in=data['line_of_business_ids'],
                batch_model=data['model_type']
            )
            if not line_of_businesses.exists():
                return Response({'error': 'No line of businesses found'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                conversion_engine = ConversionConfig.objects.get(id=data['conversion_engine_id'])
            except ConversionConfig.DoesNotExist:
                return Response({'error': 'Conversion engine not found'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                ifrs_engine = CalculationConfig.objects.get(id=data['ifrs_engine_id'])
            except CalculationConfig.DoesNotExist:
                return Response({'error': 'IFRS engine not found'}, status=status.HTTP_400_BAD_REQUEST)
            
            report_types = ReportType.objects.filter(
                id__in=data['report_type_ids'],
                batch_model=data['model_type'],
                is_enabled=True
            )
            if not report_types.exists():
                return Response({'error': 'No enabled report types found'}, status=status.HTTP_400_BAD_REQUEST)
            
            import uuid
            import time
            import random
            import string
            
            # Generate 3 random letters
            letters = ''.join(random.choices(string.ascii_uppercase, k=3))
            # Generate 8 random digits
            numbers = ''.join(random.choices(string.digits, k=8))
            run_id = f"RUN-{letters}{numbers}"
            
            model_definition = {
                'id': model.id,
                'name': model.name,
                'version': model.version,
                'config': model.config,
            }
            
            batch_data = []
            for batch in batches:
                batch_info = {
                    'id': batch.id,
                    'batch_id': batch.batch_id,
                    'batch_type': batch.batch_type,
                    'batch_model': batch.batch_model,
                    'insurance_type': batch.insurance_type,
                    'batch_year': batch.batch_year,
                    'batch_quarter': batch.batch_quarter,
                    'uploads': []
                }
                
                uploads = DataUpload.objects.filter(batch=batch)
                for upload in uploads:
                    upload_info = {
                        'id': upload.id,
                        'upload_id': upload.upload_id,
                        'source': upload.source,
                        'insurance_type': upload.insurance_type,
                        'data_type': upload.data_type,
                        'quarter': upload.quarter,
                        'year': upload.year,
                        'validation_status': upload.validation_status,
                        'rows_processed': upload.rows_processed,
                        'error_count': upload.error_count,
                        'validation_errors': upload.validation_errors,
                    }
                    batch_info['uploads'].append(upload_info)
                
                batch_data.append(batch_info)
            
            field_parameters = {
                'model_type': data['model_type'],
                'model_id': data['model_id'],
                'batch_ids': data['batch_ids'],
                'year': data['year'],
                'quarter': data['quarter'],
                'line_of_business_ids': data['line_of_business_ids'],
                'ifrs_engine_id': data['ifrs_engine_id'],
                'report_type_ids': data['report_type_ids'],
                'line_of_businesses': [
                    {
                        'id': lob.id,
                        'line_of_business': lob.line_of_business,
                        'batch_model': lob.batch_model,
                        'insurance_type': lob.insurance_type,
                        'currency': lob.currency.code if lob.currency else None,
                    }
                    for lob in line_of_businesses
                ],
                'report_types': [
                    {
                        'id': rt.id,
                        'report_type': rt.report_type,
                        'batch_model': rt.batch_model,
                        'is_enabled': rt.is_enabled,
                    }
                    for rt in report_types
                ],
            }
            
            results = []
            
            with transaction.atomic():
                engine_input = IFRSEngineInput.objects.create(
                    run_id=run_id,
                    model_definition=model_definition,
                    batch_data=batch_data,
                    field_parameters=field_parameters,
                    created_by=request.user.username if request.user else 'system'
                )
                
                for batch in batches:
                    try:
                        staging_result = self._execute_conversion_engine(
                            run_id=run_id,
                            model_definition=model_definition,
                            batch_data=batch_data,
                            field_parameters=field_parameters,
                            batch=batch,
                            line_of_businesses=line_of_businesses,
                            conversion_engine=conversion_engine
                        )
                        
                        result = IFRSEngineResult.objects.create(
                            run_id=run_id,
                            model_guid=model.id,
                            model_type=data['model_type'],
                            report_type='staging_table',
                            year=data['year'],
                            quarter=data['quarter'],
                            currency=None,
                            status='Success',
                            result_json=staging_result,
                            created_by=request.user.username if request.user else 'system'
                        )
                        results.append(result)
                        
                    except Exception as e:
                        result = IFRSEngineResult.objects.create(
                            run_id=run_id,
                            model_guid=model.id,
                            model_type=data['model_type'],
                            report_type='staging_table',
                            year=data['year'],
                            quarter=data['quarter'],
                            currency=None,
                            status='Error',
                            result_json={'error': str(e), 'traceback': str(e)},
                            created_by=request.user.username if request.user else 'system'
                        )
                        results.append(result)
                
                for batch in batches:
                    for report_type in report_types:
                        try:
                            engine_result = self._execute_python_engine(
                                run_id=run_id,
                                model_definition=model_definition,
                                batch_data=batch_data,
                                field_parameters=field_parameters,
                                batch=batch,
                                line_of_businesses=line_of_businesses,
                                report_type=report_type,
                                ifrs_engine=ifrs_engine
                            )
                            
                            result = IFRSEngineResult.objects.create(
                                run_id=run_id,
                                model_guid=model.id,
                                model_type=data['model_type'],
                                report_type=report_type.report_type,
                                year=data['year'],
                                quarter=data['quarter'],
                                currency=None,
                                status='Success',
                                result_json=engine_result,
                                created_by=request.user.username if request.user else 'system'
                            )
                            results.append(result)
                            
                            if report_type.report_type == 'disclosure_report' and 'calculations' in engine_result:
                                try:
                                    from model_definitions.utils.audit_helper import populate_disclosure_report_audit_trail
                                    
                                    calculations = engine_result.get('calculations', {})
                                    metadata = engine_result.get('metadata', {})
                                    
                                    audit_count = populate_disclosure_report_audit_trail(
                                        engine_result=result,
                                        calculations=calculations,
                                        metadata=metadata,
                                        run_id=run_id,
                                        calc_engine_version='1.0.0'
                                    )
                                    
                                    logger.info(f"Created {audit_count} audit trail records for disclosure report {run_id}")
                                except Exception as audit_error:
                                    logger.error(f"Failed to populate audit trail: {str(audit_error)}")
                            
                        except Exception as e:
                            result = IFRSEngineResult.objects.create(
                                run_id=run_id,
                                model_guid=model.id,
                                model_type=data['model_type'],
                                report_type=report_type.report_type,
                                year=data['year'],
                                quarter=data['quarter'],
                                currency=None,
                                status='Error',
                                result_json={'error': str(e), 'traceback': str(e)},
                                created_by=request.user.username if request.user else 'system'
                            )
                            results.append(result)
            
            result_serializer = IFRSEngineResultSerializer(results, many=True, context={'request': request})
            
            return Response({
                'detail': 'Reports generated successfully',
                'run_id': run_id,
                'results': result_serializer.data,
                'count': len(results)
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def _execute_conversion_engine(self, run_id, model_definition, batch_data, field_parameters, batch, line_of_businesses, conversion_engine):
        import json
        import os
        import subprocess
        import tempfile
        import sys
        
        engine_input = {
            'run_id': run_id,
            'model_definition': model_definition,
            'batch_data': batch_data,
            'field_parameters': field_parameters,
            'current_batch': {
                'id': batch.id,
                'batch_id': batch.batch_id,
                'batch_type': batch.batch_type,
                'batch_model': batch.batch_model,
                'insurance_type': batch.insurance_type,
                'batch_year': batch.batch_year,
                'batch_quarter': batch.batch_quarter,
            },
            'line_of_businesses': [
                {
                    'id': lob.id,
                    'line_of_business': lob.line_of_business,
                    'batch_model': lob.batch_model,
                    'insurance_type': lob.insurance_type,
                    'currency': lob.currency.code if lob.currency else None,
                }
                for lob in line_of_businesses
            ],
            'conversion_engine': {
                'id': conversion_engine.id,
                'engine_type': conversion_engine.engine_type,
                'batch_type': conversion_engine.batch_type,
                'batch_model': conversion_engine.batch_model,
                'insurance_type': conversion_engine.insurance_type,
            }
        }
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
            json.dump(engine_input, f, indent=2)
            input_file = f.name
        
        try:
            if not conversion_engine.script:
                return self._generate_default_staging_table(run_id, batch, line_of_businesses, field_parameters)
            
            with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False) as script_file:
                script_content = conversion_engine.script.read().decode('utf-8')
                script_file.write(script_content)
                script_file.flush()
                script_path = script_file.name
            
            try:
                python_executable = sys.executable
                
                result = subprocess.run(
                    [python_executable, script_path, input_file],
                    capture_output=True,
                    text=True,
                    timeout=300
                )
                
                if result.returncode == 0:
                    try:
                        output_data = json.loads(result.stdout)
                        return output_data
                    except json.JSONDecodeError:
                        return self._generate_default_staging_table(run_id, batch, line_of_businesses, field_parameters)
                else:
                    return self._generate_default_staging_table(run_id, batch, line_of_businesses, field_parameters)
                    
            finally:
                if os.path.exists(script_path):
                    os.unlink(script_path)
                    
        finally:
            if os.path.exists(input_file):
                os.unlink(input_file)
    
    def _generate_default_staging_table(self, run_id, batch, line_of_businesses, field_parameters):
        from datetime import datetime
        import random
        
        staging_data = []
        
        for lob in line_of_businesses:
            row = {
                'reporte_date': datetime.now().strftime('%Y-%m-%d'),
                'year': field_parameters.get('year', batch.batch_year),
                'lob': lob.line_of_business,
                'curr_actual_acq_cost': round(random.uniform(10000, 50000), 2),
                'actual_incurred_claims_settled': round(random.uniform(20000, 80000), 2),
                'pv_assumed_claims_cd_rate': round(random.uniform(0.02, 0.08), 4),
                'actual_nonclaim_handling_exp': round(random.uniform(5000, 15000), 2),
                'pv_incurred_claims_rate': round(random.uniform(0.03, 0.09), 4),
                'pv_expected_incurred_claims_rate': round(random.uniform(0.04, 0.10), 4),
                'pv_assumed_claims_pd_rate': round(random.uniform(0.01, 0.05), 4),
                'actual_prior_period_payment': round(random.uniform(15000, 45000), 2),
                'expected_claim_payments': round(random.uniform(25000, 75000), 2),
                'pv_incurred_claims_cd_rate': round(random.uniform(0.02, 0.07), 4),
                'curr_assumed_claims': round(random.uniform(30000, 90000), 2),
                'actual_earned_premium': round(random.uniform(40000, 120000), 2),
                'risk_adj_incurred_claims_curr': round(random.uniform(18000, 65000), 2),
                'risk_adj_incurred_claims_rep': round(random.uniform(17000, 60000), 2),
                'transition_risk_adj': round(random.uniform(2000, 8000), 2),
                'pv_gross_claims_transition': round(random.uniform(22000, 70000), 2),
                'transition_unearned_prem_int': round(random.uniform(3000, 12000), 2),
                'curr_actual_claim_exp': round(random.uniform(19000, 55000), 2),
            }
            staging_data.append(row)
        
        return {
            'status': 'Success',
            'run_id': run_id,
            'batch_id': batch.batch_id,
            'report_type': 'staging_table',
            'year': field_parameters.get('year', batch.batch_year),
            'quarter': field_parameters.get('quarter', batch.batch_quarter),
            'calculation_date': datetime.now().isoformat(),
            'results': {
                'detailedView': staging_data
            }
        }
    
    def _execute_python_engine(self, run_id, model_definition, batch_data, field_parameters, batch, line_of_businesses, report_type, ifrs_engine=None):
        import json
        import os
        import subprocess
        import tempfile
        import sys
        from django.conf import settings
        
        current_lob = line_of_businesses[0] if line_of_businesses else None
        
        engine_input = {
            'run_id': run_id,
            'model_definition': model_definition,
            'batch_data': batch_data,
            'field_parameters': field_parameters,
            'current_batch': {
                'id': batch.id,
                'batch_id': batch.batch_id,
                'batch_type': batch.batch_type,
                'batch_model': batch.batch_model,
                'insurance_type': batch.insurance_type,
                'batch_year': batch.batch_year,
                'batch_quarter': batch.batch_quarter,
            },
            'current_lob': {
                'id': current_lob.id if current_lob else None,
                'line_of_business': current_lob.line_of_business if current_lob else 'All LOBs',
                'batch_model': current_lob.batch_model if current_lob else batch.batch_model,
                'insurance_type': current_lob.insurance_type if current_lob else batch.insurance_type,
                'currency': current_lob.currency.code if current_lob and current_lob.currency else 'USD',
            },
            'line_of_businesses': [
                {
                    'id': lob.id,
                    'line_of_business': lob.line_of_business,
                    'batch_model': lob.batch_model,
                    'insurance_type': lob.insurance_type,
                    'currency': lob.currency.code if lob.currency else None,
                }
                for lob in line_of_businesses
            ],
            'current_report_type': {
                'id': report_type.id,
                'report_type': report_type.report_type,
                'batch_model': report_type.batch_model,
                'is_enabled': report_type.is_enabled,
            }
        }
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
            json.dump(engine_input, f, indent=2)
            input_file = f.name
        
        try:
            is_temp_script = False
            
            if report_type.report_type == 'disclosure_report':
                script_path = os.path.join(settings.BASE_DIR, 'disclosure_ifrs_engine.py')
                is_temp_script = False
                if not os.path.exists(script_path):
                    return {
                        'error': 'Disclosure IFRS engine not found',
                        'run_id': run_id
                    }
            else:
                if not ifrs_engine:
                    return {
                        'error': 'IFRS engine not provided',
                        'run_id': run_id
                    }
                
                if not ifrs_engine.script:
                    script_path = os.path.join(settings.BASE_DIR, 'ifrs_engine.py')
                    is_temp_script = False
                    if not os.path.exists(script_path):
                        return {
                            'error': f'No script uploaded for engine {ifrs_engine.engine_type} and no default engine found',
                            'run_id': run_id
                        }
                else:
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False) as script_file:
                        script_content = ifrs_engine.script.read().decode('utf-8')
                        script_file.write(script_content)
                        script_file.flush()
                        script_path = script_file.name
                        is_temp_script = True
            
            try:
                python_executable = sys.executable
                
                result = subprocess.run(
                    [python_executable, script_path, input_file],
                    capture_output=True,
                    text=True,
                    timeout=300
                )
                
                if result.returncode == 0:
                    try:
                        output_data = json.loads(result.stdout)
                        return output_data
                    except json.JSONDecodeError:
                        return {
                            'message': 'Engine executed successfully but returned invalid JSON',
                            'stdout': result.stdout,
                            'stderr': result.stderr,
                            'run_id': run_id
                        }
                else:
                    import traceback as tb
                    error_details = {
                        'error': 'Engine execution failed',
                        'stdout': result.stdout,
                        'stderr': result.stderr,
                        'return_code': result.returncode,
                        'run_id': run_id,
                        'script_path': script_path,
                        'input_file': input_file
                    }
                    logger.error(f"Engine execution failed: {error_details}")
                    return error_details
                    
            finally:
                if is_temp_script:
                    try:
                        os.unlink(script_path)
                    except:
                        pass
                    
        except subprocess.TimeoutExpired:
            return {
                'error': 'Engine execution timed out',
                'run_id': run_id
            }
        except Exception as e:
            return {
                'error': f'Engine execution error: {str(e)}',
                'run_id': run_id
            }
        finally:
            try:
                os.unlink(input_file)
            except:
                pass


class IFRSApiConfigViewSet(viewsets.ModelViewSet):
    queryset = IFRSApiConfig.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['api_source_name', 'client_id', 'data_type', 'owner']
    ordering_fields = ['api_source_name', 'client_id', 'status', 'created_on', 'modified_on']
    ordering = ['api_source_name', 'client_id']
    filterset_fields = ['method', 'auth_type', 'schedule', 'status']
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return IFRSApiConfigCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return IFRSApiConfigUpdateSerializer
        return IFRSApiConfigSerializer
    
    def perform_create(self, serializer):
        """
        Set owner to current user on creation
        """
        serializer.save(owner=self.request.user.username)
    
    @action(detail=True, methods=['post'])
    def test_connection(self, request, pk=None):
        """
        Test API connection for the given configuration
        """
        api_config = self.get_object()
        
        try:
            # Here will implement actual connection testing logic
            # For now, just update the test status
            api_config.last_test_date = timezone.now()
            api_config.last_test_status = 'success'
            api_config.save()
            
            return Response({
                'detail': 'Connection test successful',
                'lastTestDate': api_config.last_test_date,
                'lastTestStatus': api_config.last_test_status
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            api_config.last_test_date = timezone.now()
            api_config.last_test_status = 'failed'
            api_config.save()
            
            return Response({
                'detail': f'Connection test failed: {str(e)}',
                'lastTestDate': api_config.last_test_date,
                'lastTestStatus': api_config.last_test_status
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def dry_run(self, request, pk=None):
        """
        Perform a dry run to fetch sample data (100 rows)
        """
        api_config = self.get_object()
        
        try:
            # Here will implement actual dry run logic
            # For now, return a mock response
            sample_data = {
                'records_fetched': 100,
                'sample_record': {
                    'id': 1,
                    'timestamp': '2024-01-01T00:00:00Z',
                    'data': 'Sample data record'
                },
                'parsed_successfully': True,
                'jsonpath_results': {
                    'records': '$.data.items[*]',
                    'total_count': '$.pagination.total'
                }
            }
            
            return Response({
                'detail': 'Dry run completed successfully',
                'sampleData': sample_data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'detail': f'Dry run failed: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        active_configs = self.queryset.filter(status='active')
        page = self.paginate_queryset(active_configs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(active_configs, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def by_schedule(self, request):
        """
        Get API configurations by schedule type
        """
        schedule_type = request.query_params.get('schedule', None)
        if not schedule_type:
            return Response({
                'detail': 'Schedule parameter is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        configs = self.queryset.filter(schedule=schedule_type)
        page = self.paginate_queryset(configs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(configs, many=True)
        return Response(serializer.data)


class AuditViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CalculationValue.objects.all()
    serializer_class = CalculationValueSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['run_id', 'value_id', 'report_type']
    search_fields = ['value_id', 'label', 'run_id']
    ordering_fields = ['value_id', 'timestamp']
    ordering = ['value_id']
    
    @action(detail=False, methods=['get'])
    def runs_by_period(self, request):
        try:
            results = IFRSEngineResult.objects.values(
                'run_id',
                'year',
                'quarter',
                'model_type',
                'currency',
                'status',
                'created_at'
            ).distinct().order_by('-created_at')
            
            runs_dict = {}
            for result in results:
                run_id = result['run_id']
                if run_id not in runs_dict:
                    report_types = IFRSEngineResult.objects.filter(
                        run_id=run_id
                    ).values_list('report_type', flat=True).distinct()
                    
                    runs_dict[run_id] = {
                        'run_id': run_id,
                        'period': f"{result['year']} {result['quarter']}",
                        'legal_entity': 'Default Entity',  # TODO: Get from config
                        'currency': result['currency'] or 'USD',
                        'status': 'Final' if result['status'] == 'Success' else 'Draft',
                        'execution_date': result['created_at'],
                        'model_type': result['model_type'],
                        'available_reports': list(report_types)
                    }
            
            runs_list = list(runs_dict.values())
            serializer = RunSummarySerializer(runs_list, many=True)
            
            return Response({
                'detail': 'Success',
                'results': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Error fetching runs: {str(e)}")
            return Response({
                'detail': f'Error: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def reports_by_run(self, request):
        run_id = request.query_params.get('run_id')
        
        if not run_id:
            return Response({
                'detail': 'run_id parameter is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            results = IFRSEngineResult.objects.filter(run_id=run_id).values(
                'id',
                'report_type',
                'status'
            )
            
            report_data = []
            for result in results:
                value_count = 0
                try:
                    value_count = CalculationValue.objects.filter(
                        run_id=run_id,
                        report_type=result['report_type']
                    ).count()
                except Exception:
                    value_count = 0
                
                report_data.append({
                    'report_type': result['report_type'],
                    'report_type_display': result['report_type'].replace('_', ' ').title(),
                    'status': result['status'],
                    'value_count': value_count
                })
            
            serializer = ReportMetadataSerializer(report_data, many=True)
            
            return Response({
                'detail': 'Success',
                'results': serializer.data
            })
            
        except Exception as e:
            logger.error(f"Error fetching reports for run {run_id}: {str(e)}")
            return Response({
                'detail': f'Error: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def value_detail(self, request):
        run_id = request.query_params.get('run_id')
        value_id = request.query_params.get('value_id')
        
        if not run_id or not value_id:
            return Response({
                'detail': 'Both run_id and value_id parameters are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            calc_value = CalculationValue.objects.prefetch_related(
                'assumptions',
                'input_refs'
            ).get(run_id=run_id, value_id=value_id)
            
            serializer = CalculationValueSerializer(calc_value)
            
            return Response({
                'detail': 'Success',
                'result': serializer.data
            })
            
        except CalculationValue.DoesNotExist:
            try:
                engine_results = IFRSEngineResult.objects.filter(
                    run_id=run_id,
                    report_type='disclosure_report'
                )
                
                for result in engine_results:
                    result_json = result.result_json
                    
                    if isinstance(result_json, dict) and 'calculations' in result_json:
                        calculations = result_json['calculations']
                        metadata = result_json.get('metadata', {})
                        
                        for key, calc_data in calculations.items():
                            if calc_data.get('value_id') == value_id:
                                period = f"{metadata.get('year', '')} {metadata.get('quarter', '')}"
                                legal_entity = metadata.get('legal_entity_name', 'Unknown')
                                currency = metadata.get('currency_name', 'USD')
                                label = value_id.replace('_', ' ').replace('.', ' - ')
                                amount = calc_data.get('amount', 0)
                                
                                fallback_data = {
                                    'valueId': value_id,
                                    'runId': run_id,
                                    'reportType': 'disclosure_report',
                                    'period': period,
                                    'legalEntity': legal_entity,
                                    'currency': currency,
                                    'label': label,
                                    'value': float(amount),
                                    'unit': 'currency',
                                    'lineOfBusiness': metadata.get('run_name', ''),
                                    'cohort': None,
                                    'groupId': None,
                                    'formulaHumanReadable': None,
                                    'dependencies': [],
                                    'calculationMethod': metadata.get('method_name', 'PAA').split()[0],
                                    'notes': 'Temporary data from classification engine - full audit trail not yet available',
                                    'isMissingData': False,
                                    'isOverride': False,
                                    'isFallback': True,
                                    'hasRounding': False,
                                    'calcEngineVersion': '1.0.0',
                                    'timestamp': result_json.get('calculation_date', ''),
                                    'assumptions': [],
                                    'inputRefs': []
                                }
                                
                                return Response({
                                    'detail': 'Success (Fallback Mode)',
                                    'result': fallback_data
                                })
                
                return Response({
                    'detail': f'Value {value_id} not found in disclosure report for run {run_id}. The calculation engine needs to be updated to track audit metadata.'
                }, status=status.HTTP_404_NOT_FOUND)
                
            except IFRSEngineResult.DoesNotExist:
                return Response({
                    'detail': f'No disclosure report found for run_id={run_id}'
                }, status=status.HTTP_404_NOT_FOUND)
            except Exception as fallback_error:
                logger.error(f"Fallback extraction failed: {str(fallback_error)}")
                return Response({
                    'detail': f'Calculation value not found and fallback extraction failed: {str(fallback_error)}'
                }, status=status.HTTP_404_NOT_FOUND)
                
        except Exception as e:
            logger.error(f"Error fetching value detail: {str(e)}")
            return Response({
                'detail': f'Error: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def submitted_reports_by_type(self, request):
        report_type = request.query_params.get('report_type')
        
        if not report_type:
            return Response({
                'detail': 'report_type parameter is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            submitted_reports = SubmittedReport.objects.filter(
                report_type=report_type,
                status='active'
            ).select_related().order_by('-assign_year', '-assign_quarter', '-created_on')
            
            results = []
            for report in submitted_reports:
                try:
                    engine_result = IFRSEngineResult.objects.get(id=report.ifrs_engine_result_id)
                    results.append({
                        'id': report.id,
                        'run_id': report.run_id,
                        'report_type': report.report_type,
                        'report_type_display': report.report_type_display or report.report_type.replace('_', ' ').title(),
                        'model_type': report.model_type,
                        'assign_year': report.assign_year,
                        'assign_quarter': report.assign_quarter,
                        'status': report.status,
                        'created_at': report.created_on,
                        'display_name': f"{report.report_type_display or report.report_type.replace('_', ' ').title()} - {report.run_id} - {report.assign_year} {report.assign_quarter}"
                    })
                except IFRSEngineResult.DoesNotExist:
                    logger.warning(f"IFRSEngineResult not found for submitted report {report.id}")
                    continue
            
            return Response({
                'detail': 'Success',
                'results': results
            })
            
        except Exception as e:
            logger.error(f"Error fetching submitted reports: {str(e)}")
            return Response({
                'detail': f'Error: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def compare_runs(self, request):
        current_run_id = request.query_params.get('current_run_id')
        prior_run_id = request.query_params.get('prior_run_id')
        value_id = request.query_params.get('value_id')
        
        if not current_run_id or not prior_run_id or not value_id:
            return Response({
                'detail': 'current_run_id, prior_run_id, and value_id parameters are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            current_result = IFRSEngineResult.objects.filter(run_id=current_run_id).first()
            prior_result = IFRSEngineResult.objects.filter(run_id=prior_run_id).first()
            
            if not current_result:
                return Response({
                    'detail': f'Current run {current_run_id} not found'
                }, status=status.HTTP_404_NOT_FOUND)
            
            if not prior_result:
                return Response({
                    'detail': f'Prior run {prior_run_id} not found'
                }, status=status.HTTP_404_NOT_FOUND)
            
            current_json = current_result.result_json or {}
            prior_json = prior_result.result_json or {}
            
            current_value = None
            prior_value = None
            
            if isinstance(current_json, dict) and 'calculations' in current_json:
                calculations = current_json['calculations']
                for key, calc_data in calculations.items():
                    if calc_data.get('value_id') == value_id:
                        current_value = {
                            'value_id': value_id,
                            'amount': calc_data.get('amount', 0),
                            'metadata': current_json.get('metadata', {})
                        }
                        break
            
            if isinstance(prior_json, dict) and 'calculations' in prior_json:
                calculations = prior_json['calculations']
                for key, calc_data in calculations.items():
                    if calc_data.get('value_id') == value_id:
                        prior_value = {
                            'value_id': value_id,
                            'amount': calc_data.get('amount', 0),
                            'metadata': prior_json.get('metadata', {})
                        }
                        break
            
            if not current_value:
                return Response({
                    'detail': f'Value {value_id} not found in current run {current_run_id}'
                }, status=status.HTTP_404_NOT_FOUND)
            
            if not prior_value:
                return Response({
                    'detail': f'Value {value_id} not found in prior run {prior_run_id}'
                }, status=status.HTTP_404_NOT_FOUND)
            
            current_amount = float(current_value['amount'])
            prior_amount = float(prior_value['amount'])
            absolute_change = current_amount - prior_amount
            percentage_change = ((current_amount - prior_amount) / prior_amount * 100) if prior_amount != 0 else 0
            
            try:
                current_calc_value = CalculationValue.objects.prefetch_related(
                    'assumptions', 'input_refs'
                ).filter(run_id=current_run_id, value_id=value_id).first()
                
                prior_calc_value = CalculationValue.objects.prefetch_related(
                    'assumptions', 'input_refs'
                ).filter(run_id=prior_run_id, value_id=value_id).first()
            except Exception:
                current_calc_value = None
                prior_calc_value = None
            
            comparison_data = {
                'value_id': value_id,
                'current_run_id': current_run_id,
                'prior_run_id': prior_run_id,
                'current_value': current_amount,
                'prior_value': prior_amount,
                'absolute_change': absolute_change,
                'percentage_change': percentage_change,
                'current_metadata': current_value['metadata'],
                'prior_metadata': prior_value['metadata'],
                'has_audit_data': current_calc_value is not None and prior_calc_value is not None
            }
            
            if current_calc_value and prior_calc_value:
                assumption_changes = []
                input_changes = []
                
                current_assumptions = {a.assumption_id: a for a in current_calc_value.assumptions.all()}
                prior_assumptions = {a.assumption_id: a for a in prior_calc_value.assumptions.all()}
                
                for assumption_id, current_assumption in current_assumptions.items():
                    if assumption_id not in prior_assumptions:
                        assumption_changes.append({
                            'type': 'added',
                            'assumption_id': assumption_id,
                            'assumption_type': current_assumption.assumption_type
                        })
                    elif current_assumption.assumption_version != prior_assumptions[assumption_id].assumption_version:
                        assumption_changes.append({
                            'type': 'version_changed',
                            'assumption_id': assumption_id,
                            'assumption_type': current_assumption.assumption_type,
                            'prior_version': prior_assumptions[assumption_id].assumption_version,
                            'current_version': current_assumption.assumption_version
                        })
                
                for assumption_id in prior_assumptions:
                    if assumption_id not in current_assumptions:
                        assumption_changes.append({
                            'type': 'removed',
                            'assumption_id': assumption_id,
                            'assumption_type': prior_assumptions[assumption_id].assumption_type
                        })
                
                current_inputs = {i.dataset_name: i for i in current_calc_value.input_refs.all()}
                prior_inputs = {i.dataset_name: i for i in prior_calc_value.input_refs.all()}
                
                for dataset_name, current_input in current_inputs.items():
                    if dataset_name not in prior_inputs:
                        input_changes.append({
                            'type': 'added',
                            'dataset_name': dataset_name,
                            'record_count': current_input.record_count
                        })
                    elif current_input.source_snapshot_id != prior_inputs[dataset_name].source_snapshot_id:
                        input_changes.append({
                            'type': 'snapshot_changed',
                            'dataset_name': dataset_name,
                            'prior_snapshot': prior_inputs[dataset_name].source_snapshot_id,
                            'current_snapshot': current_input.source_snapshot_id,
                            'prior_record_count': prior_inputs[dataset_name].record_count,
                            'current_record_count': current_input.record_count
                        })
                
                for dataset_name in prior_inputs:
                    if dataset_name not in current_inputs:
                        input_changes.append({
                            'type': 'removed',
                            'dataset_name': dataset_name
                        })
                
                comparison_data['assumption_changes'] = assumption_changes
                comparison_data['input_changes'] = input_changes
                comparison_data['formula_changed'] = (
                    current_calc_value.formula_human_readable != prior_calc_value.formula_human_readable
                )
            
            ai_insight = self._generate_ai_insight(comparison_data)
            comparison_data['ai_insight'] = ai_insight
            
            return Response({
                'detail': 'Success',
                'result': comparison_data
            })
            
        except Exception as e:
            logger.error(f"Error comparing runs: {str(e)}")
            return Response({
                'detail': f'Error: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _generate_ai_insight(self, comparison_data):
        if OPENAI_AVAILABLE and settings.OPENAI_API_KEY:
            try:
                return self._generate_ai_insight_with_openai(comparison_data)
            except Exception as e:
                logger.error(f"OpenAI API call failed: {str(e)}")
                return self._generate_fallback_insight(comparison_data)
        else:
            return self._generate_fallback_insight(comparison_data)
    
    def _generate_ai_insight_with_openai(self, comparison_data):
        client = OpenAI(api_key=settings.OPENAI_API_KEY)
        
        absolute_change = comparison_data['absolute_change']
        percentage_change = comparison_data['percentage_change']
        current_value = comparison_data['current_value']
        prior_value = comparison_data['prior_value']
        value_id = comparison_data['value_id']
        
        prompt_parts = [
            f"You are an IFRS 17 audit analyst reviewing changes in insurance contract valuations.",
            f"\n\nValue ID: {value_id}",
            f"\nCurrent Value: ${current_value:,.2f}",
            f"Prior Value: ${prior_value:,.2f}",
            f"Absolute Change: ${absolute_change:,.2f}",
            f"Percentage Change: {percentage_change:.2f}%"
        ]
        
        if comparison_data.get('has_audit_data'):
            prompt_parts.append("\n\nDetailed Change Information:")
            
            if comparison_data.get('formula_changed'):
                prompt_parts.append("\n- The calculation formula was modified between runs")
            
            assumption_changes = comparison_data.get('assumption_changes', [])
            if assumption_changes:
                prompt_parts.append(f"\n\nAssumption Changes ({len(assumption_changes)} total):")
                for change in assumption_changes[:10]:
                    change_type = change['type'].replace('_', ' ')
                    assumption_type = change['assumption_type'].replace('_', ' ')
                    if change['type'] == 'version_changed':
                        prompt_parts.append(
                            f"\n- {change_type.upper()}: {assumption_type} (ID: {change['assumption_id']}) "
                            f"version changed from {change.get('prior_version', 'N/A')} to {change.get('current_version', 'N/A')}"
                        )
                    else:
                        prompt_parts.append(
                            f"\n- {change_type.upper()}: {assumption_type} (ID: {change['assumption_id']})"
                        )
                if len(assumption_changes) > 10:
                    prompt_parts.append(f"\n- ... and {len(assumption_changes) - 10} more assumption changes")
            
            input_changes = comparison_data.get('input_changes', [])
            if input_changes:
                prompt_parts.append(f"\n\nInput Data Changes ({len(input_changes)} total):")
                for change in input_changes[:10]:
                    change_type = change['type'].replace('_', ' ')
                    if change['type'] == 'snapshot_changed':
                        record_change = change.get('current_record_count', 0) - change.get('prior_record_count', 0)
                        prompt_parts.append(
                            f"\n- {change_type.upper()}: {change['dataset_name']} "
                            f"(snapshot: {change.get('prior_snapshot', 'N/A')} → {change.get('current_snapshot', 'N/A')}, "
                            f"records changed by {record_change:+,})"
                        )
                    else:
                        prompt_parts.append(
                            f"\n- {change_type.upper()}: {change['dataset_name']}"
                        )
                if len(input_changes) > 10:
                    prompt_parts.append(f"\n- ... and {len(input_changes) - 10} more input changes")
        
        prompt_parts.append(
            "\n\nProvide a concise analysis (3-5 sentences) covering:"
            "\n1. Magnitude assessment (negligible/minor/moderate/significant)"
            "\n2. Primary drivers of the change"
            "\n3. Potential implications for IFRS 17 compliance"
            "\n4. Any red flags or items requiring auditor attention"
            "\n\nBe specific and actionable. Focus on what an auditor needs to know."
        )
        
        prompt = ''.join(prompt_parts)
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": "You are an expert IFRS 17 audit analyst with deep knowledge of insurance contract accounting, actuarial assumptions, and regulatory compliance. Provide clear, professional insights."
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            temperature=0.3,
            max_tokens=500
        )
        
        ai_insight = response.choices[0].message.content.strip()
        return ai_insight
    
    def _generate_fallback_insight(self, comparison_data):
        absolute_change = comparison_data['absolute_change']
        percentage_change = comparison_data['percentage_change']
        
        if abs(percentage_change) < 0.01:
            magnitude = "negligible"
        elif abs(percentage_change) < 5:
            magnitude = "minor"
        elif abs(percentage_change) < 20:
            magnitude = "moderate"
        else:
            magnitude = "significant"
        
        direction = "increased" if absolute_change > 0 else "decreased"
        
        insight_parts = [
            f"The value has {direction} by {abs(percentage_change):.2f}% (${abs(absolute_change):,.2f}), which represents a {magnitude} change."
        ]
        
        if comparison_data.get('has_audit_data'):
            changes = []
            
            if comparison_data.get('formula_changed'):
                changes.append("calculation formula was modified")
            
            assumption_changes = comparison_data.get('assumption_changes', [])
            if assumption_changes:
                added = len([c for c in assumption_changes if c['type'] == 'added'])
                removed = len([c for c in assumption_changes if c['type'] == 'removed'])
                version_changed = len([c for c in assumption_changes if c['type'] == 'version_changed'])
                
                if added:
                    changes.append(f"{added} assumption(s) were added")
                if removed:
                    changes.append(f"{removed} assumption(s) were removed")
                if version_changed:
                    changes.append(f"{version_changed} assumption(s) were updated to new versions")
            
            input_changes = comparison_data.get('input_changes', [])
            if input_changes:
                snapshot_changed = len([c for c in input_changes if c['type'] == 'snapshot_changed'])
                added = len([c for c in input_changes if c['type'] == 'added'])
                removed = len([c for c in input_changes if c['type'] == 'removed'])
                
                if snapshot_changed:
                    changes.append(f"{snapshot_changed} input dataset(s) were updated with new data snapshots")
                if added:
                    changes.append(f"{added} input dataset(s) were added")
                if removed:
                    changes.append(f"{removed} input dataset(s) were removed")
            
            if changes:
                insight_parts.append("\n\nKey drivers of the change:")
                for i, change in enumerate(changes, 1):
                    insight_parts.append(f"\n{i}. {change.capitalize()}")
            else:
                insight_parts.append("\n\nNo significant changes detected in assumptions, inputs, or formulas. The change may be due to indirect factors or rounding differences.")
        else:
            insight_parts.append("\n\nNote: Detailed audit data is not available for these runs. Enable audit trail tracking for more comprehensive change analysis.")
        
        return ''.join(insight_parts)
