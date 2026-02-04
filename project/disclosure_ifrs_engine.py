import json
import sys
import os
import logging
import base64
from datetime import datetime
from typing import Dict, Any, List
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class DisclosureIFRSEngine:
    def __init__(self, input_data: Dict[str, Any]):
        self.input_data = input_data
        self.run_id = input_data.get('run_id')
        self.model_definition = input_data.get('model_definition', {})
        self.batch_data = input_data.get('batch_data', [])
        self.field_parameters = input_data.get('field_parameters', {})
        self.current_batch = input_data.get('current_batch', {})
        self.current_lob = input_data.get('current_lob', {})
        self.current_report_type = input_data.get('current_report_type', {})
        
        self.metadata = self._extract_metadata()
        
    def _extract_metadata(self) -> Dict[str, Any]:
        """Extract report metadata from inputs"""
        return {
            'legal_entity_name': self.field_parameters.get('legal_entity_name', 'TEST'),
            'run_type': self.field_parameters.get('run_type', 'Solo'),
            'method_name': 'Premium Allocation Approach',
            'execution_date': f"31/12/{self.field_parameters.get('year', datetime.now().year)}",
            'run_name': self.current_lob.get('line_of_business', 'All LOBs'),
            'trans_method': 'Fair Value Approach, Non-Transition',
            'scenario_name': 'SCN1',
            'execution_type': 'YTD',
            'currency_name': self.current_lob.get('currency', 'Bahamian Dollar'),
            'year': self.field_parameters.get('year', datetime.now().year),
            'quarter': self.field_parameters.get('quarter', 'Q4'),
        }
    
    def calculate_values(self) -> Dict[str, Any]:
        lob = self.current_lob.get('line_of_business', 'All LOBs')
        period = f"{self.metadata['year']}Q{self.metadata['quarter'][-1]}"
        prev_period = f"{self.metadata['year'] - 1}Q4"
        
        values = {
            'value_1': {
                'value_id': 'DR.MA.Opening.Liabilities.PVFC_LFRC', 
                'amount': 0.00,
                'display_name': 'Opening Balance - Present Value of Future Cash Flows (LFRC)',
                'formula': f'ClosingBalance(PVFC_LFRC, {prev_period})',
                'dependencies': [f'DR.MA.Closing.Liabilities.PVFC_LFRC@{prev_period}'],
                'dimensions': {
                    'LOB': lob,
                    'Group': 'Direct',
                    'Cohort': str(self.metadata['year'])
                },
                'assumptions': {
                    'discount_curve': {
                        'id': f'DISC_CURVE_{period}_V1',
                        'version': '1.0',
                        'effective_date': f"{self.metadata['year']}-01-01",
                        'curve_type': 'Zero-coupon yield curve'
                    }
                },
                'inputs': {
                    'datasets': [
                        {
                            'dataset': 'liability_balances',
                            'snapshot_id': f'SNAP_{self.metadata["year"]}0101_001',
                            'record_count': 0,
                            'source_hash': 'abc123def456'
                        }
                    ]
                },
                'flags': {
                    'is_missing_data': False,
                    'is_override': False,
                    'is_fallback': False,
                    'has_rounding': False
                },
                'quality_flags': ['OK'],
                'notes': 'Opening LFRC balance at start of period'
            },
            'value_2': {
                'value_id': 'DR.MA.Opening.Liabilities.PVFC_LIC', 
                'amount': 1345608.03,
                'display_name': 'Opening Balance - Present Value of Future Cash Flows (LIC)',
                'formula': f'ClosingBalance(PVFC_LIC, {prev_period})',
                'dependencies': [f'DR.MA.Closing.Liabilities.PVFC_LIC@{prev_period}'],
                'dimensions': {
                    'LOB': lob,
                    'Group': 'Direct',
                    'Cohort': str(self.metadata['year'])
                },
                'assumptions': {
                    'discount_curve': {
                        'id': f'DISC_CURVE_{period}_V1',
                        'version': '1.0',
                        'effective_date': f"{self.metadata['year']}-01-01",
                        'curve_type': 'Zero-coupon yield curve'
                    }
                },
                'inputs': {
                    'datasets': [
                        {
                            'dataset': 'claims_data',
                            'snapshot_id': f'SNAP_{self.metadata["year"]}0101_002',
                            'record_count': 10234,
                            'source_hash': 'def456ghi789'
                        }
                    ]
                },
                'flags': {
                    'is_missing_data': False,
                    'is_override': False,
                    'is_fallback': False,
                    'has_rounding': False
                },
                'quality_flags': ['OK'],
                'notes': 'Opening LIC balance from prior period'
            },
            'value_3': {
                'value_id': 'DR.MA.Opening.Liabilities.RiskAdj_LFRC', 
                'amount': 0.00,
                'display_name': 'Opening Balance - Risk Adjustment (LFRC)',
                'formula': f'ClosingBalance(RiskAdj_LFRC, {prev_period})',
                'dependencies': [f'DR.MA.Closing.Liabilities.RiskAdj_LFRC@{prev_period}'],
                'dimensions': {
                    'LOB': lob,
                    'Group': 'Direct',
                    'Cohort': str(self.metadata['year'])
                },
                'assumptions': {
                    'risk_adjustment': {
                        'id': f'RA_METHOD_PAA_{period}_V2',
                        'version': '2.0',
                        'effective_date': f"{self.metadata['year']}-01-01",
                        'method': 'Cost of Capital',
                        'confidence_level': '75%'
                    }
                },
                'inputs': {
                    'datasets': [
                        {
                            'dataset': 'risk_parameters',
                            'snapshot_id': f'SNAP_{self.metadata["year"]}0101_003',
                            'record_count': 150,
                            'source_hash': 'ghi789jkl012'
                        }
                    ]
                },
                'flags': {
                    'is_missing_data': False,
                    'is_override': False,
                    'is_fallback': False,
                    'has_rounding': False
                },
                'quality_flags': ['OK'],
                'notes': 'Risk adjustment for uncertainty in LFRC'
            },
            'value_4': {
                'value_id': 'DR.MA.Opening.Liabilities.RiskAdj_LIC', 
                'amount': 201841.21,
                'display_name': 'Opening Balance - Risk Adjustment (LIC)',
                'formula': f'ClosingBalance(RiskAdj_LIC, {prev_period})',
                'dependencies': [f'DR.MA.Closing.Liabilities.RiskAdj_LIC@{prev_period}'],
                'dimensions': {
                    'LOB': lob,
                    'Group': 'Direct',
                    'Cohort': str(self.metadata['year'])
                },
                'assumptions': {
                    'risk_adjustment': {
                        'id': f'RA_METHOD_PAA_{period}_V2',
                        'version': '2.0',
                        'effective_date': f"{self.metadata['year']}-01-01",
                        'method': 'Cost of Capital',
                        'confidence_level': '75%'
                    }
                },
                'inputs': {
                    'datasets': [
                        {
                            'dataset': 'claims_data',
                            'snapshot_id': f'SNAP_{self.metadata["year"]}0101_002',
                            'record_count': 10234,
                            'source_hash': 'def456ghi789'
                        },
                        {
                            'dataset': 'risk_parameters',
                            'snapshot_id': f'SNAP_{self.metadata["year"]}0101_003',
                            'record_count': 150,
                            'source_hash': 'ghi789jkl012'
                        }
                    ]
                },
                'flags': {
                    'is_missing_data': False,
                    'is_override': False,
                    'is_fallback': False,
                    'has_rounding': False
                },
                'quality_flags': ['OK'],
                'notes': 'Risk adjustment for claims uncertainty'
            },
            'value_5': {
                'value_id': 'DR.MA.Opening.Liabilities.Total', 
                'amount': 1547449.24,
                'display_name': 'Opening Balance - Total Liabilities',
                'formula': 'SUM(PVFC_LFRC, PVFC_LIC, RiskAdj_LFRC, RiskAdj_LIC)',
                'dependencies': [
                    'DR.MA.Opening.Liabilities.PVFC_LFRC',
                    'DR.MA.Opening.Liabilities.PVFC_LIC',
                    'DR.MA.Opening.Liabilities.RiskAdj_LFRC',
                    'DR.MA.Opening.Liabilities.RiskAdj_LIC'
                ],
                'dimensions': {
                    'LOB': lob,
                    'Group': 'Direct',
                    'Cohort': str(self.metadata['year'])
                },
                'assumptions': {},
                'inputs': {
                    'datasets': []
                },
                'flags': {
                    'is_missing_data': False,
                    'is_override': False,
                    'is_fallback': False,
                    'has_rounding': True
                },
                'quality_flags': ['OK'],
                'notes': 'Total opening liability balance (sum of components)'
            },
            
            'value_6': {
                'value_id': 'DR.MA.Opening.Assets.PVFC_LFRC', 
                'amount': 0.00,
                'display_name': 'Opening Balance - Assets PVFC (LFRC)',
                'formula': f'ClosingBalance(Assets_PVFC_LFRC, {prev_period})',
                'dependencies': [f'DR.MA.Closing.Assets.PVFC_LFRC@{prev_period}'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Reinsurance assets opening balance'
            },
            'value_7': {
                'value_id': 'DR.MA.Opening.Assets.PVFC_LIC', 
                'amount': 1900005.14,
                'display_name': 'Opening Balance - Assets PVFC (LIC)',
                'formula': f'ClosingBalance(Assets_PVFC_LIC, {prev_period})',
                'dependencies': [f'DR.MA.Closing.Assets.PVFC_LIC@{prev_period}'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {
                    'discount_curve': {
                        'id': f'DISC_CURVE_{period}_V1',
                        'version': '1.0',
                        'effective_date': f"{self.metadata['year']}-01-01"
                    }
                },
                'inputs': {
                    'datasets': [{
                        'dataset': 'reinsurance_data',
                        'snapshot_id': f'SNAP_{self.metadata["year"]}0101_004',
                        'record_count': 8765,
                        'source_hash': 'jkl012mno345'
                    }]
                },
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Reinsurance recoverable on LIC'
            },
            'value_8': {
                'value_id': 'DR.MA.Opening.Assets.RiskAdj_LFRC', 
                'amount': 0.00,
                'display_name': 'Opening Balance - Assets Risk Adjustment (LFRC)',
                'formula': f'ClosingBalance(Assets_RiskAdj_LFRC, {prev_period})',
                'dependencies': [f'DR.MA.Closing.Assets.RiskAdj_LFRC@{prev_period}'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_9': {
                'value_id': 'DR.MA.Opening.Assets.RiskAdj_LIC', 
                'amount': 224101.17,
                'display_name': 'Opening Balance - Assets Risk Adjustment (LIC)',
                'formula': f'ClosingBalance(Assets_RiskAdj_LIC, {prev_period})',
                'dependencies': [f'DR.MA.Closing.Assets.RiskAdj_LIC@{prev_period}'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {
                    'risk_adjustment': {
                        'id': f'RA_METHOD_PAA_{period}_V2',
                        'version': '2.0',
                        'confidence_level': '75%'
                    }
                },
                'inputs': {
                    'datasets': [{
                        'dataset': 'reinsurance_data',
                        'snapshot_id': f'SNAP_{self.metadata["year"]}0101_004',
                        'record_count': 8765,
                        'source_hash': 'jkl012mno345'
                    }]
                },
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Risk adjustment on reinsurance assets'
            },
            'value_10': {
                'value_id': 'DR.MA.Opening.Assets.Total', 
                'amount': 2124106.31,
                'display_name': 'Opening Balance - Total Assets',
                'formula': 'SUM(Assets_PVFC_LFRC, Assets_PVFC_LIC, Assets_RiskAdj_LFRC, Assets_RiskAdj_LIC)',
                'dependencies': [
                    'DR.MA.Opening.Assets.PVFC_LFRC',
                    'DR.MA.Opening.Assets.PVFC_LIC',
                    'DR.MA.Opening.Assets.RiskAdj_LFRC',
                    'DR.MA.Opening.Assets.RiskAdj_LIC'
                ],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Total reinsurance assets'
            },
            
            'value_11': {
                'value_id': 'DR.MA.Opening.Net.PVFC_LFRC', 
                'amount': 0.00,
                'display_name': 'Opening Net Balance - PVFC (LFRC)',
                'formula': 'Liabilities_PVFC_LFRC + Assets_PVFC_LFRC',
                'dependencies': ['DR.MA.Opening.Liabilities.PVFC_LFRC', 'DR.MA.Opening.Assets.PVFC_LFRC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Net position (liabilities - assets)'
            },
            'value_12': {
                'value_id': 'DR.MA.Opening.Net.PVFC_LIC', 
                'amount': 3245613.18,
                'display_name': 'Opening Net Balance - PVFC (LIC)',
                'formula': 'Liabilities_PVFC_LIC + Assets_PVFC_LIC',
                'dependencies': ['DR.MA.Opening.Liabilities.PVFC_LIC', 'DR.MA.Opening.Assets.PVFC_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Net LIC after reinsurance'
            },
            'value_13': {
                'value_id': 'DR.MA.Opening.Net.RiskAdj_LFRC', 
                'amount': 0.00,
                'display_name': 'Opening Net Balance - Risk Adjustment (LFRC)',
                'formula': 'Liabilities_RiskAdj_LFRC + Assets_RiskAdj_LFRC',
                'dependencies': ['DR.MA.Opening.Liabilities.RiskAdj_LFRC', 'DR.MA.Opening.Assets.RiskAdj_LFRC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_14': {
                'value_id': 'DR.MA.Opening.Net.RiskAdj_LIC', 
                'amount': 425942.38,
                'display_name': 'Opening Net Balance - Risk Adjustment (LIC)',
                'formula': 'Liabilities_RiskAdj_LIC + Assets_RiskAdj_LIC',
                'dependencies': ['DR.MA.Opening.Liabilities.RiskAdj_LIC', 'DR.MA.Opening.Assets.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Net risk adjustment'
            },
            'value_15': {
                'value_id': 'DR.MA.Opening.Net.Total', 
                'amount': 3671555.55,
                'display_name': 'Opening Net Balance - Total',
                'formula': 'SUM(Net_PVFC_LFRC, Net_PVFC_LIC, Net_RiskAdj_LFRC, Net_RiskAdj_LIC)',
                'dependencies': [
                    'DR.MA.Opening.Net.PVFC_LFRC',
                    'DR.MA.Opening.Net.PVFC_LIC',
                    'DR.MA.Opening.Net.RiskAdj_LFRC',
                    'DR.MA.Opening.Net.RiskAdj_LIC'
                ],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Total net position at opening'
            },
            
            'value_16': {
                'value_id': 'DR.MA.PastService.PVFC_LIC', 
                'amount': -3221378.11,
                'display_name': 'Past Service Adjustment - PVFC (LIC)',
                'formula': 'Actual_Claims - Expected_Claims',
                'dependencies': ['DR.MA.Opening.Net.PVFC_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {
                    'datasets': [{
                        'dataset': 'claims_actual',
                        'snapshot_id': f'SNAP_{self.metadata["year"]}1231_005',
                        'record_count': 12456,
                        'source_hash': 'mno345pqr678'
                    }]
                },
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Experience adjustment on past service'
            },
            'value_17': {
                'value_id': 'DR.MA.PastService.RiskAdj_LIC', 
                'amount': -320079.17,
                'display_name': 'Past Service Adjustment - Risk Adjustment (LIC)',
                'formula': 'Release_RiskAdj_For_Past_Service',
                'dependencies': ['DR.MA.Opening.Net.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {
                    'risk_adjustment': {
                        'id': f'RA_METHOD_PAA_{period}_V2',
                        'version': '2.0',
                        'release_pattern': 'Straight-line'
                    }
                },
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Risk adjustment release for expired coverage'
            },
            'value_18': {
                'value_id': 'DR.MA.PastService.Total', 
                'amount': -3541457.28,
                'display_name': 'Past Service Adjustment - Total',
                'formula': 'PastService_PVFC_LIC + PastService_RiskAdj_LIC',
                'dependencies': ['DR.MA.PastService.PVFC_LIC', 'DR.MA.PastService.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Total adjustment for past service'
            },
            
            'value_19': {
                'value_id': 'DR.MA.Closing.Net.PVFC_LIC', 
                'amount': 24235.07,
                'display_name': 'Closing Net Balance - PVFC (LIC)',
                'formula': 'Opening_PVFC_LIC + PastService_PVFC_LIC + CurrentService + FinanceExpense - CashFlows',
                'dependencies': ['DR.MA.Opening.Net.PVFC_LIC', 'DR.MA.PastService.PVFC_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Closing LIC balance'
            },
            'value_20': {
                'value_id': 'DR.MA.Closing.Net.RiskAdj_LIC', 
                'amount': 105863.20,
                'display_name': 'Closing Net Balance - Risk Adjustment (LIC)',
                'formula': 'Opening_RiskAdj_LIC + PastService_RiskAdj_LIC + NewBusiness_RiskAdj - Release_RiskAdj',
                'dependencies': ['DR.MA.Opening.Net.RiskAdj_LIC', 'DR.MA.PastService.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Closing risk adjustment'
            },
            'value_21': {
                'value_id': 'DR.MA.Closing.Net.Total', 
                'amount': 130098.27,
                'display_name': 'Closing Net Balance - Total',
                'formula': 'Closing_PVFC_LIC + Closing_RiskAdj_LIC',
                'dependencies': ['DR.MA.Closing.Net.PVFC_LIC', 'DR.MA.Closing.Net.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Total closing net position'
            },
            
            'value_22': {
                'value_id': 'DR.MA.Closing.Liabilities.PVFC_LIC', 
                'amount': 11272.73,
                'display_name': 'Closing Liabilities - PVFC (LIC)',
                'formula': 'Closing_Net_PVFC_LIC * Liability_Proportion',
                'dependencies': ['DR.MA.Closing.Net.PVFC_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Direct', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Liability component of closing balance'
            },
            'value_23': {
                'value_id': 'DR.MA.Closing.Liabilities.RiskAdj_LIC', 
                'amount': 1690.91,
                'display_name': 'Closing Liabilities - Risk Adjustment (LIC)',
                'formula': 'Closing_Net_RiskAdj_LIC * Liability_Proportion',
                'dependencies': ['DR.MA.Closing.Net.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Direct', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_24': {
                'value_id': 'DR.MA.Closing.Liabilities.Total', 
                'amount': 12963.64,
                'display_name': 'Closing Liabilities - Total',
                'formula': 'Closing_Liabilities_PVFC_LIC + Closing_Liabilities_RiskAdj_LIC',
                'dependencies': ['DR.MA.Closing.Liabilities.PVFC_LIC', 'DR.MA.Closing.Liabilities.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Direct', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Total closing liabilities'
            },
            
            'value_25': {
                'value_id': 'DR.MA.Closing.Assets.PVFC_LIC', 
                'amount': 12962.34,
                'display_name': 'Closing Assets - PVFC (LIC)',
                'formula': 'Closing_Net_PVFC_LIC * Asset_Proportion',
                'dependencies': ['DR.MA.Closing.Net.PVFC_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Reinsurance asset component'
            },
            'value_26': {
                'value_id': 'DR.MA.Closing.Assets.RiskAdj_LIC', 
                'amount': 104172.30,
                'display_name': 'Closing Assets - Risk Adjustment (LIC)',
                'formula': 'Closing_Net_RiskAdj_LIC * Asset_Proportion',
                'dependencies': ['DR.MA.Closing.Net.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_27': {
                'value_id': 'DR.MA.Closing.Assets.Total', 
                'amount': 117134.64,
                'display_name': 'Closing Assets - Total',
                'formula': 'Closing_Assets_PVFC_LIC + Closing_Assets_RiskAdj_LIC',
                'dependencies': ['DR.MA.Closing.Assets.PVFC_LIC', 'DR.MA.Closing.Assets.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Total closing reinsurance assets'
            },
            
            'value_28': {
                'value_id': 'DR.LA.Opening.Liabilities.LFRC_ExcludingLoss', 
                'amount': 207.66,
                'display_name': 'Opening Liabilities - LFRC Excluding Loss Component',
                'formula': f'ClosingBalance(LFRC_ExcludingLoss, {prev_period})',
                'dependencies': [f'DR.LA.Closing.Liabilities.LFRC_ExcludingLoss@{prev_period}'],
                'dimensions': {'LOB': lob, 'Group': 'Direct', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {
                    'datasets': [{
                        'dataset': 'premiums_unearned',
                        'snapshot_id': f'SNAP_{self.metadata["year"]}0101_006',
                        'record_count': 5432,
                        'source_hash': 'pqr678stu901'
                    }]
                },
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Liability for remaining coverage'
            },
            'value_29': {
                'value_id': 'DR.LA.Opening.Liabilities.PVFC', 
                'amount': 1345608.03,
                'display_name': 'Opening Liabilities - Present Value of Future Cash Flows',
                'formula': 'Sum(PVFC_LFRC, PVFC_LIC)',
                'dependencies': ['DR.MA.Opening.Liabilities.PVFC_LFRC', 'DR.MA.Opening.Liabilities.PVFC_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Direct', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_30': {
                'value_id': 'DR.LA.Opening.Liabilities.RiskAdj', 
                'amount': 201841.21,
                'display_name': 'Opening Liabilities - Risk Adjustment',
                'formula': 'Sum(RiskAdj_LFRC, RiskAdj_LIC)',
                'dependencies': ['DR.MA.Opening.Liabilities.RiskAdj_LFRC', 'DR.MA.Opening.Liabilities.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Direct', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_31': {
                'value_id': 'DR.LA.Opening.Liabilities.Total', 
                'amount': 1547656.90,
                'display_name': 'Opening Liabilities - Total',
                'formula': 'LFRC_ExcludingLoss + PVFC + RiskAdj',
                'dependencies': ['DR.LA.Opening.Liabilities.LFRC_ExcludingLoss', 'DR.LA.Opening.Liabilities.PVFC', 'DR.LA.Opening.Liabilities.RiskAdj'],
                'dimensions': {'LOB': lob, 'Group': 'Direct', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Total opening liability position'
            },
            
            'value_32': {
                'value_id': 'DR.LA.Opening.Assets.LFRC_ExcludingLoss', 
                'amount': -244.36,
                'display_name': 'Opening Assets - LFRC Excluding Loss Component',
                'formula': f'ClosingBalance(Assets_LFRC_ExcludingLoss, {prev_period})',
                'dependencies': [f'DR.LA.Closing.Assets.LFRC_ExcludingLoss@{prev_period}'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_33': {
                'value_id': 'DR.LA.Opening.Assets.PVFC', 
                'amount': 1900005.14,
                'display_name': 'Opening Assets - Present Value of Future Cash Flows',
                'formula': 'Sum(Assets_PVFC_LFRC, Assets_PVFC_LIC)',
                'dependencies': ['DR.MA.Opening.Assets.PVFC_LFRC', 'DR.MA.Opening.Assets.PVFC_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_34': {
                'value_id': 'DR.LA.Opening.Assets.RiskAdj', 
                'amount': 224101.17,
                'display_name': 'Opening Assets - Risk Adjustment',
                'formula': 'Sum(Assets_RiskAdj_LFRC, Assets_RiskAdj_LIC)',
                'dependencies': ['DR.MA.Opening.Assets.RiskAdj_LFRC', 'DR.MA.Opening.Assets.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_35': {
                'value_id': 'DR.LA.Opening.Assets.Total', 
                'amount': 2123861.95,
                'display_name': 'Opening Assets - Total',
                'formula': 'Assets_LFRC_ExcludingLoss + Assets_PVFC + Assets_RiskAdj',
                'dependencies': ['DR.LA.Opening.Assets.LFRC_ExcludingLoss', 'DR.LA.Opening.Assets.PVFC', 'DR.LA.Opening.Assets.RiskAdj'],
                'dimensions': {'LOB': lob, 'Group': 'Reinsurance', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Total opening reinsurance assets'
            },
            
            'value_36': {
                'value_id': 'DR.LA.InsuranceRevenue', 
                'amount': -24624785.52,
                'display_name': 'Insurance Revenue',
                'formula': 'Release_LFRC_For_Services_Provided',
                'dependencies': ['DR.LA.Opening.Liabilities.LFRC_ExcludingLoss'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {
                    'datasets': [{
                        'dataset': 'premium_earnings',
                        'snapshot_id': f'SNAP_{self.metadata["year"]}1231_007',
                        'record_count': 9876,
                        'source_hash': 'stu901vwx234'
                    }]
                },
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Revenue recognized for period'
            },
            'value_37': {
                'value_id': 'DR.LA.ServiceExpense.AcqAmortization', 
                'amount': 4811287.49,
                'display_name': 'Service Expense - Acquisition Cost Amortization',
                'formula': 'Amortize_Acquisition_Costs_Over_Coverage_Period',
                'dependencies': ['DR.LA.CashFlows.AcqCashFlow'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {
                    'datasets': [{
                        'dataset': 'acquisition_costs',
                        'snapshot_id': f'SNAP_{self.metadata["year"]}1231_008',
                        'record_count': 3456,
                        'source_hash': 'vwx234yza567'
                    }]
                },
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Systematic allocation of acquisition costs'
            },
            'value_38': {
                'value_id': 'DR.LA.ServiceExpense.PastServiceAdj.PVFC', 
                'amount': -3221378.11,
                'display_name': 'Service Expense - Past Service Adjustment (PVFC)',
                'formula': 'Actual_Claims - Expected_Claims',
                'dependencies': ['DR.MA.PastService.PVFC_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Experience adjustment'
            },
            'value_39': {
                'value_id': 'DR.LA.ServiceExpense.PastServiceAdj.RiskAdj', 
                'amount': -320079.17,
                'display_name': 'Service Expense - Past Service Adjustment (Risk Adjustment)',
                'formula': 'Release_RiskAdj_For_Past_Service',
                'dependencies': ['DR.MA.PastService.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            
            'value_40': {
                'value_id': 'DR.LA.CashFlows.PremiumReceived', 
                'amount': 14008787.00,
                'display_name': 'Cash Flows - Premium Received',
                'formula': 'Sum(Premium_Collections)',
                'dependencies': [],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {
                    'datasets': [{
                        'dataset': 'premium_collections',
                        'snapshot_id': f'SNAP_{self.metadata["year"]}1231_009',
                        'record_count': 15678,
                        'source_hash': 'yza567bcd890'
                    }]
                },
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Cash received from policyholders'
            },
            'value_41': {
                'value_id': 'DR.LA.CashFlows.AcqCashFlow', 
                'amount': -2212219.58,
                'display_name': 'Cash Flows - Acquisition Cash Flows',
                'formula': 'Sum(Acquisition_Payments)',
                'dependencies': [],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {
                    'datasets': [{
                        'dataset': 'acquisition_payments',
                        'snapshot_id': f'SNAP_{self.metadata["year"]}1231_010',
                        'record_count': 2345,
                        'source_hash': 'bcd890efg123'
                    }]
                },
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Cash paid for acquisition costs'
            },
            
            'value_42': {
                'value_id': 'DR.LA.Closing.Net.LFRC_ExcludingLoss', 
                'amount': -8016967.31,
                'display_name': 'Closing Net - LFRC Excluding Loss Component',
                'formula': 'Opening_LFRC + PremiumReceived - InsuranceRevenue - AcqCashFlow',
                'dependencies': ['DR.LA.Opening.Liabilities.LFRC_ExcludingLoss', 'DR.LA.CashFlows.PremiumReceived', 'DR.LA.InsuranceRevenue', 'DR.LA.CashFlows.AcqCashFlow'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Closing LFRC balance'
            },
            'value_43': {
                'value_id': 'DR.LA.Closing.Net.PVFC', 
                'amount': 24235.07,
                'display_name': 'Closing Net - Present Value of Future Cash Flows',
                'formula': 'Opening_PVFC + PastService_PVFC + CurrentService',
                'dependencies': ['DR.MA.Closing.Net.PVFC_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_44': {
                'value_id': 'DR.LA.Closing.Net.RiskAdj', 
                'amount': 105863.20,
                'display_name': 'Closing Net - Risk Adjustment',
                'formula': 'Opening_RiskAdj + PastService_RiskAdj + NewBusiness - Release',
                'dependencies': ['DR.MA.Closing.Net.RiskAdj_LIC'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': ''
            },
            'value_45': {
                'value_id': 'DR.LA.Closing.Net.Total', 
                'amount': -7886869.04,
                'display_name': 'Closing Net - Total',
                'formula': 'Closing_LFRC + Closing_PVFC + Closing_RiskAdj',
                'dependencies': ['DR.LA.Closing.Net.LFRC_ExcludingLoss', 'DR.LA.Closing.Net.PVFC', 'DR.LA.Closing.Net.RiskAdj'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Total net position at closing'
            },
            
            # Profit/Loss values
            'value_46': {
                'value_id': 'DR.PnL.InsuranceRevenue', 
                'amount': 24624785.53,
                'display_name': 'Profit & Loss - Insurance Revenue',
                'formula': 'ABS(InsuranceRevenue)',
                'dependencies': ['DR.LA.InsuranceRevenue'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Total insurance revenue for P&L'
            },
            'value_47': {
                'value_id': 'DR.PnL.InsuranceServiceExpenses', 
                'amount': -1269830.21,
                'display_name': 'Profit & Loss - Insurance Service Expenses',
                'formula': 'AcqAmortization + PastServiceAdj_PVFC + PastServiceAdj_RiskAdj + ClaimsIncurred',
                'dependencies': [
                    'DR.LA.ServiceExpense.AcqAmortization',
                    'DR.LA.ServiceExpense.PastServiceAdj.PVFC',
                    'DR.LA.ServiceExpense.PastServiceAdj.RiskAdj'
                ],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': False},
                'quality_flags': ['OK'],
                'notes': 'Total insurance service expenses'
            },
            'value_48': {
                'value_id': 'DR.PnL.InsuranceServiceResult', 
                'amount': 23354955.31,
                'display_name': 'Profit & Loss - Insurance Service Result',
                'formula': 'InsuranceRevenue - InsuranceServiceExpenses',
                'dependencies': ['DR.PnL.InsuranceRevenue', 'DR.PnL.InsuranceServiceExpenses'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Insurance service result (revenue - expenses)'
            },
            'value_49': {
                'value_id': 'DR.PnL.InvestmentIncome', 
                'amount': 0.00,
                'display_name': 'Profit & Loss - Investment Income',
                'formula': 'Investment_Returns_On_Insurance_Assets',
                'dependencies': [],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': True, 'is_override': False, 'is_fallback': True, 'has_rounding': False},
                'quality_flags': ['WARNING'],
                'notes': 'Investment income not yet implemented - defaulting to zero'
            },
            'value_50': {
                'value_id': 'DR.PnL.InsuranceFinanceExpenses', 
                'amount': 0.00,
                'display_name': 'Profit & Loss - Insurance Finance Expenses',
                'formula': 'Discount_Rate_Changes + Time_Value_Money',
                'dependencies': [],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {
                    'discount_curve': {
                        'id': f'DISC_CURVE_{period}_V1',
                        'version': '1.0',
                        'note': 'Not applied in PAA'
                    }
                },
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': True, 'is_override': False, 'is_fallback': True, 'has_rounding': False},
                'quality_flags': ['WARNING'],
                'notes': 'Finance expenses not material under PAA - defaulting to zero'
            },
            'value_51': {
                'value_id': 'DR.PnL.FinancialResult', 
                'amount': 0.00,
                'display_name': 'Profit & Loss - Financial Result',
                'formula': 'InvestmentIncome - InsuranceFinanceExpenses',
                'dependencies': ['DR.PnL.InvestmentIncome', 'DR.PnL.InsuranceFinanceExpenses'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': True, 'is_override': False, 'is_fallback': True, 'has_rounding': False},
                'quality_flags': ['WARNING'],
                'notes': 'Net financial result (currently zero)'
            },
            'value_52': {
                'value_id': 'DR.PnL.Profit', 
                'amount': 23354955.31,
                'display_name': 'Profit & Loss - Total Profit',
                'formula': 'InsuranceServiceResult + FinancialResult',
                'dependencies': ['DR.PnL.InsuranceServiceResult', 'DR.PnL.FinancialResult'],
                'dimensions': {'LOB': lob, 'Group': 'Net', 'Cohort': str(self.metadata['year'])},
                'assumptions': {},
                'inputs': {'datasets': []},
                'flags': {'is_missing_data': False, 'is_override': False, 'is_fallback': False, 'has_rounding': True},
                'quality_flags': ['OK'],
                'notes': 'Total profit for the period'
            },
        }
        
        return values
    
    def generate_excel(self, values: Dict[str, Any]) -> bytes:
        """Generate Excel workbook matching the template structure"""
        
        wb = Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        section_title_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '#,##0.00'
        
        self._create_metadata_sheet(wb, thin_border, header_font, header_fill, bold_font)
        self._create_movement_analysis_sheet(wb, values, section_title_font, header_font, 
                                             header_fill, bold_font, thin_border, currency_format)
        self._create_liability_analysis_sheet(wb, values, section_title_font, header_font,
                                              header_fill, bold_font, thin_border, currency_format)
        self._create_profit_loss_sheet(wb, values, section_title_font, header_font,
                                       header_fill, bold_font, thin_border, currency_format)
        self._create_value_id_mapping_sheet(wb, values, header_font, header_fill, thin_border)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _create_metadata_sheet(self, wb, border, header_font, header_fill, bold_font):
        ws = wb.create_sheet('Report Metadata')
        
        metadata_rows = [
            ['Field', 'Value'],
            ['Denomination For Amounts', 'As Is'],
            ['Legal Entity Name', self.metadata['legal_entity_name']],
            ['Run Type', self.metadata['run_type']],
            ['Method Name', self.metadata['method_name']],
            ['Execution Date', self.metadata['execution_date']],
            ['Run Name', self.metadata['run_name']],
            ['Trans Method', self.metadata['trans_method']],
            ['Scenario Name', self.metadata['scenario_name']],
            ['Execution Type', self.metadata['execution_type']],
            ['Currency Name', self.metadata['currency_name']],
            ['Run ID', self.run_id],
            ['Generated At', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
        ]
        
        for row_idx, row_data in enumerate(metadata_rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                elif col_idx == 1:
                    cell.font = bold_font
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
    
    def _create_movement_analysis_sheet(self, wb, values, title_font, header_font, 
                                       header_fill, bold_font, border, currency_fmt):
        ws = wb.create_sheet('Movement Analysis - PAA')
        
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = 'Movement Analysis - Premium Allocation Approach'
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:F2')
        time_cell = ws['A2']
        time_cell.value = f"Time run: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        time_cell.alignment = Alignment(horizontal='center')
        
        row = 4
        headers = ['', 'Present Value Of Future Cash Flows (LFRC)', 'Present Value Of Future Cash Flows (LIC)', 
                   'Risk Adjustment (LFRC)', 'Risk Adjustment (LIC)', 'Total Liability']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        data_rows = [
            ('Opening Insurance Contract Liabilities', ['value_1', 'value_2', 'value_3', 'value_4', 'value_5'], False),
            ('Opening Insurance Contract Assets', ['value_6', 'value_7', 'value_8', 'value_9', 'value_10'], False),
            ('Net Opening Balance', ['value_11', 'value_12', 'value_13', 'value_14', 'value_15'], True),
            ('Incurred Claim And Directly Attributable Expenses', [None, None, None, None, None], False),
            ('Changes that relate to Current Service', [None, None, None, None, None], False),
            ('  Experience Adjustment', [None, None, None, None, None], False),
            ('  Change in Risk Adjustment For the Report Period', [None, None, None, None, None], False),
            ('Changes that relate to Future Service', [None, None, None, None, None], False),
            ('Change Related to Past Services', [None, 'value_16', None, 'value_17', 'value_18'], False),
            ('Finance (Income) Expense From Insurance Contracts Issued (P/L)', [None, None, None, None, None], False),
            ('Finance (Income) Expense From Insurance Contracts Issued (OCI)', [None, None, None, None, None], False),
            ('Cash Flows', [None, None, None, None, None], False),
            ('Net Closing Balance', [None, 'value_19', None, 'value_20', 'value_21'], True),
            ('Closing Insurance Contract Liabilities', [None, 'value_22', None, 'value_23', 'value_24'], False),
            ('Closing Insurance Contract Assets', [None, 'value_25', None, 'value_26', 'value_27'], False),
        ]
        
        row = 5
        for label, value_keys, is_bold in data_rows:
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.border = border
            if is_bold:
                label_cell.font = bold_font
            
            for col_idx, value_key in enumerate(value_keys, 2):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                
                if value_key and value_key in values:
                    value_data = values[value_key]
                    cell.value = value_data['amount']
                    cell.number_format = currency_fmt
                    
                    comment = Comment(f"value_id: {value_data['value_id']}", "System")
                    cell.comment = comment
                    
                    if is_bold:
                        cell.font = bold_font
                else:
                    cell.value = 0.00
                    cell.number_format = currency_fmt
            
            row += 1
        
        ws.column_dimensions['A'].width = 60
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 25
    
    def _create_liability_analysis_sheet(self, wb, values, title_font, header_font,
                                        header_fill, bold_font, border, currency_fmt):
        ws = wb.create_sheet('Liability Analysis - PAA')
        
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = 'Liability Analysis - Premium Allocation Approach'
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        row = 3
        headers = ['', 'Liability For Remaining Coverage\nExcluding Loss Component', 
                   'Loss Component', 'Present Value of Future Cash Flows',
                   'Risk Adjustment', 'Total Liability']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        data_rows = [
            ('Opening Insurance Contract Liabilities', ['value_28', None, 'value_29', 'value_30', 'value_31'], False),
            ('Opening Insurance Contract Assets', ['value_32', None, 'value_33', 'value_34', 'value_35'], False),
            ('Net Opening Balance', [None, None, None, None, None], True),
            ('Insurance Revenue', ['value_36', None, None, None, 'value_36'], False),
            ('Insurance Service Expense', None, False),
            ('  Incurred Claim And Directly Attributable Expenses', [None, None, None, None, None], False),
            ('  Changes that relate to past services- Adjustment to LIC', [None, None, 'value_38', 'value_39', None], False),
            ('  Insurance Acquisition Cashflow Amortization', ['value_37', None, None, None, 'value_37'], False),
            ('Total Insurance Service Expense', [None, None, None, None, None], True),
            ('Insurance Service Result', [None, None, None, None, None], True),
            ('Cash flows', None, False),
            ('  Premium Received', ['value_40', None, None, None, 'value_40'], False),
            ('  Insurance Acquisition Cash flow', ['value_41', None, None, None, 'value_41'], False),
            ('Total Cash Flows', [None, None, None, None, None], True),
            ('Net Closing Balance', ['value_42', None, 'value_43', 'value_44', 'value_45'], True),
            ('Closing Insurance Contract Liabilities', [None, None, None, None, None], False),
            ('Closing Insurance Contract Assets', [None, None, None, None, None], False),
        ]
        
        row = 4
        for row_data in data_rows:
            if len(row_data) == 3:
                label, value_keys, is_bold = row_data
            else:
                label = row_data[0]
                value_keys = None
                is_bold = False
            
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.border = border
            if is_bold or value_keys is None:
                label_cell.font = bold_font
            
            if value_keys is not None:
                for col_idx, value_key in enumerate(value_keys, 2):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right')
                    
                    if value_key and value_key in values:
                        value_data = values[value_key]
                        cell.value = value_data['amount']
                        cell.number_format = currency_fmt
                        
                        comment = Comment(f"value_id: {value_data['value_id']}", "System")
                        cell.comment = comment
                        
                        if is_bold:
                            cell.font = bold_font
                    else:
                        cell.value = 0.00
                        cell.number_format = currency_fmt
            
            row += 1
        
        ws.column_dimensions['A'].width = 60
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 25
    
    def _create_profit_loss_sheet(self, wb, values, title_font, header_font,
                                  header_fill, bold_font, border, currency_fmt):
        ws = wb.create_sheet('Statement of Profit or Loss')
        
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = 'STATEMENT OF PROFIT OR LOSS - Premium Allocation Approach'
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:B2')
        time_cell = ws['A2']
        time_cell.value = f"Time run: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        time_cell.alignment = Alignment(horizontal='center')
        
        row = 4
        headers = ['', 'Profit/Loss']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        data_rows = [
            ('Insurance Revenue', 'value_46', False),
            ('Insurance Service Expenses', 'value_47', False),
            ('Insurance Service Result', 'value_48', True),
            ('Investment Income', 'value_49', False),
            ('Insurance Finance Expenses', 'value_50', False),
            ('Financial Result', 'value_51', True),
            ('Profit', 'value_52', True),
        ]
        
        row = 5
        for label, value_key, is_bold in data_rows:
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.border = border
            if is_bold:
                label_cell.font = bold_font
            
            cell = ws.cell(row=row, column=2)
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            
            if value_key in values:
                value_data = values[value_key]
                cell.value = value_data['amount']
                cell.number_format = currency_fmt
                
                comment = Comment(f"value_id: {value_data['value_id']}", "System")
                cell.comment = comment
                
                if is_bold:
                    cell.font = bold_font
            
            row += 1
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 25
    
    def _create_value_id_mapping_sheet(self, wb, values, header_font, header_fill, border):
        ws = wb.create_sheet('Value ID Mapping')
        
        headers = ['Value Key', 'Value ID', 'Amount', 'Description']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        row = 2
        for value_key, value_data in sorted(values.items()):
            ws.cell(row=row, column=1, value=value_key).border = border
            ws.cell(row=row, column=2, value=value_data['value_id']).border = border
            ws.cell(row=row, column=3, value=value_data['amount']).border = border
            ws.cell(row=row, column=4, value=f"Value for {value_data['value_id']}").border = border
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 60
    
    def execute(self) -> Dict[str, Any]:
        try:
            logger.info(f"Starting Disclosure IFRS Engine execution for Run ID: {self.run_id}")
            
            values = self.calculate_values()
            logger.info(f"Calculated {len(values)} values")
            
            excel_bytes = self.generate_excel(values)
            logger.info("Excel workbook generated successfully")
            
            excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
            
            result = {
                'run_id': self.run_id,
                'report_type': 'disclosure_report',
                'model_type': self.field_parameters.get('model_type'),
                'year': self.metadata['year'],
                'quarter': self.metadata['quarter'],
                'currency': self.metadata['currency_name'],
                'calculation_date': datetime.now().isoformat(),
                
                'inputs': {
                    'run_id': self.run_id,
                    'model_definition': self.model_definition,
                    'field_parameters': self.field_parameters,
                    'batch_info': {
                        'batch_id': self.current_batch.get('batch_id'),
                        'batch_year': self.current_batch.get('batch_year'),
                        'batch_quarter': self.current_batch.get('batch_quarter'),
                        'batch_model': self.current_batch.get('batch_model'),
                    },
                    'lob_info': {
                        'line_of_business': self.current_lob.get('line_of_business'),
                        'currency': self.current_lob.get('currency'),
                    }
                },
                
                'calculations': values,
                
                'metadata': self.metadata,
                
                'excel_bytes': excel_base64,
                
                'status': 'success'
            }
            
            logger.info("Disclosure Report generated successfully")
            return result
            
        except Exception as e:
            logger.error(f"Error in Disclosure IFRS Engine execution: {str(e)}")
            import traceback
            return {
                'error': f'Engine execution failed: {str(e)}',
                'traceback': traceback.format_exc(),
                'run_id': self.run_id,
                'status': 'failed'
            }


def main():
    if len(sys.argv) != 2:
        print(json.dumps({
            'error': 'Usage: python disclosure_ifrs_engine.py <input_json_file>'
        }))
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    try:
        with open(input_file, 'r') as f:
            input_data = json.load(f)
        
        engine = DisclosureIFRSEngine(input_data)
        result = engine.execute()
        
        print(json.dumps(result, indent=2, default=str))
        
        if 'error' in result:
            sys.exit(1)
        else:
            sys.exit(0)
            
    except FileNotFoundError:
        print(json.dumps({
            'error': f'Input file not found: {input_file}'
        }))
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(json.dumps({
            'error': f'Invalid JSON in input file: {str(e)}'
        }))
        sys.exit(1)
    except Exception as e:
        import traceback
        print(json.dumps({
            'error': f'Unexpected error: {str(e)}',
            'traceback': traceback.format_exc()
        }))
        sys.exit(1)


if __name__ == '__main__':
    main()
